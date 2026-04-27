import random
import os
from pathlib import Path
from django.db import transaction
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Protection
from openpyxl.worksheet.protection import SheetProtection
from floodsApp.models import State, City, Tehsil, Village  # Adjust import based on your app structure

# ----------------------------------------------------------------------
# 1. CONFIGURATION
# ----------------------------------------------------------------------
LOCATION_EXCEL = Path('media/excel_reports/location_data_20251016_230542.xlsx')  # For village IDs only
OUTPUT_DIR = Path('excel_reports')
OUTPUT_DIR.mkdir(exist_ok=True)
OUTPUT_FILE = OUTPUT_DIR / 'dummy_farmers_data_500_new.xlsx'

# ----------------------------------------------------------------------
# 2. LOAD VALID LOCATION HIERARCHY FROM MODELS
# ----------------------------------------------------------------------
def load_location_hierarchy():
    """Fetch valid state, city, tehsil, and village combinations from the database."""
    with transaction.atomic():  # Ensure consistent reads
        states = list(State.objects.all())
        if not states:
            raise ValueError("No states found in the database")
        
        state = states[0]  # Assuming only Punjab for now
        cities = list(City.objects.filter(state=state))
        if not cities:
            raise ValueError("No cities found for the state")
        
        village_list = []
        for city in cities:
            tehsils = list(Tehsil.objects.filter(city=city))
            if not tehsils:
                continue
            villages = list(Village.objects.filter(tehsil__in=tehsils))
            for village in villages:
                village_list.append((village.id, state.id, city.id, village.tehsil.id))
        
        if not village_list:
            raise ValueError("No valid village hierarchy found")
        
        return village_list

# Load hierarchy once
village_list = load_location_hierarchy()

# ----------------------------------------------------------------------
# 3. DUMMY DATA LISTS (names, statuses, etc.)
# ----------------------------------------------------------------------
first_names = [
    'Gurpreet', 'Jaswinder', 'Baljit', 'Amarjeet', 'Manpreet',
    'Harpreet', 'Sukhwinder', 'Daljit', 'Karan', 'Ravinder',
    'Simran', 'Ranjit', 'Kuldeep', 'Sandeep', 'Navjot',
    'Hardeep', 'Paramjit', 'Jagjit', 'Satnam', 'Ajay',
    'Pardeep', 'Rajinder', 'Mandeep', 'Surinder', 'Balwinder',
    'Gagan', 'Harpal', 'Jaspal', 'Kamaljit', 'Lakhwinder',
    'Narinder', 'Prabhjot', 'Rajwinder', 'Sukhdev', 'Taranjit',
    'Amritpal', 'Charanjit', 'Darshan', 'Gurmail', 'Harjinder',
    'Jagmohan', 'Kashmir', 'Kirpal', 'Manjit', 'Nirmal',
    'Raghbir', 'Satpal', 'Shamsher', 'Surjit', 'Tejinder'
]
last_names = ['Singh', 'Kaur', 'Sharma']
father_names = [
    'Harbans Singh', 'Gurcharan Singh', 'Baldev Singh',
    'Sohan Singh', 'Amrik Singh'
]
surveyor_names = [
    'Amarjeet Singh', 'Manpreet Kaur', 'Ranjit Singh',
    'Simerpreet Kaur', 'Kuldeep Singh'
]
house_statuses = ['Partially Damaged', 'No Damage', 'Fully Damaged']
crop_losses = ['10,000 - 25,000', '25,000 - 50,000', '50,000 - 100,000']
yes_no_options = ['Yes', 'No', 'Not Required']
land_values = ['5', '10', '15', '20']
crop_values = ['50', '100', '200', '300']

# ----------------------------------------------------------------------
# 4. HEADERS (must match the bulk-upload view)
# ----------------------------------------------------------------------
headers = [
    'farmerName', 'fatherName', 'mobileNumber', 'email', 'state_id',
    'city_id', 'tehsil_id', 'village_id', 'pincode', 'houseStatus',
    'totalLandOwned', 'landAffected', 'cropsPlanted', 'cropsLost',
    'estimatedCropLoss', 'tractorLeveling', 'manureFertilizer',
    'seedsRequired', 'fertilizersPesticides', 'laborRequirement',
    'irrigationRepair', 'livestockDamage', 'householdNeeds',
    'housingRepair', 'otherSupport', 'surveyorName', 'AmountNeeded'
]

# ----------------------------------------------------------------------
# 5. BUILD 500 ROWS, cycling through all village_ids
# ----------------------------------------------------------------------
def generate_row(village_idx: int):
    """Create one row using the village at `village_idx` (mod len(village_list))."""
    village_id, state_id, city_id, tehsil_id = village_list[village_idx % len(village_list)]

    first = random.choice(first_names)
    last = random.choice(last_names)
    farmer_name = f"{first} {last}"
    email = f"{first.lower()}.{last.lower()}{random.randint(100, 999)}@gmail.com" if random.random() > 0.3 else ''

    total_land = random.choice(land_values)
    affected_land = str(random.randint(1, int(float(total_land))))
    planted = random.choice(crop_values)
    lost = str(random.randint(10, int(float(planted))))
    amount_needed = random.randint(5000, 50000)

    return [
        farmer_name,
        random.choice(father_names),
        f"{random.randint(7,9)}{random.randint(10000000,99999999)}",
        email,
        state_id,
        city_id,
        tehsil_id,
        village_id,
        '',  # pincode – left empty
        random.choice(house_statuses),
        total_land,
        affected_land,
        planted,
        lost,
        random.choice(crop_losses),
        random.choice(yes_no_options),
        random.choice(yes_no_options),
        random.choice(yes_no_options),
        random.choice(yes_no_options),
        random.choice(yes_no_options),
        random.choice(yes_no_options),
        random.choice(yes_no_options),
        random.choice(yes_no_options),
        random.choice(yes_no_options),
        random.choice(yes_no_options),
        random.choice(surveyor_names),
        amount_needed
    ]

# Generate the data (500 rows)
rows = []
for i in range(500):
    rows.append(generate_row(i))

# ----------------------------------------------------------------------
# 6. CREATE THE EXCEL FILE
# ----------------------------------------------------------------------
wb = Workbook()
ws = wb.active
ws.title = 'Farmers Data'

# ---- Headers -------------------------------------------------------
for col, hdr in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=hdr)
    cell.font = Font(bold=True)
    cell.protection = Protection(locked=True)

# ---- Data ---------------------------------------------------------
for r_idx, row_data in enumerate(rows, start=2):
    for c_idx, value in enumerate(row_data, start=1):
        cell = ws.cell(row=r_idx, column=c_idx, value=value)
        cell.protection = Protection(locked=False)

# ---- Auto-fit columns -----------------------------------------------
for column in ws.columns:
    max_len = 0
    col_letter = column[0].column_letter
    for cell in column:
        try:
            if len(str(cell.value)) > max_len:
                max_len = len(str(cell.value))
        except Exception:
            pass
    adjusted_width = min(max_len + 2, 50)
    ws.column_dimensions[col_letter].width = adjusted_width

# ---- Protect the sheet -----------------------------------------------
ws.protection = SheetProtection(sheet=True, objects=False, scenarios=False)

# ---- Save -----------------------------------------------------------
wb.save(OUTPUT_FILE)
print(f"Excel file with 500 rows saved to: {OUTPUT_FILE}")