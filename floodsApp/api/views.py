import uuid
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
import json
from floodsApp.models import *
from django.utils.dateparse import parse_date
from urllib.parse import urljoin
from django.db import transaction
import traceback
from rest_framework.pagination import PageNumberPagination
from django_filters.rest_framework import DjangoFilterBackend
from floodsApp.api.filters import TempPersonDataFormFilter
import os
from datetime import datetime
from django.conf import settings
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.worksheet.protection import SheetProtection
from openpyxl.styles import Protection
from openpyxl import load_workbook
from decimal import Decimal

from django.db.models import Sum
from django.core.exceptions import ObjectDoesNotExist, ValidationError

from weasyprint import HTML
from jinja2 import Template
from django.http import HttpResponse
from django.templatetags.static import static
from .persondatapdf import get_pdf_template
from collections import OrderedDict  

class TempPersonPagination(PageNumberPagination):
    page_size = 5
    page_size_query_param = 'page_size'
    max_page_size = 100  # Optional: Set a maximum page size to prevent abuse

    def get_page_size(self, request):
        page_size = request.query_params.get(self.page_size_query_param)
        if page_size:
            try:
                return min(int(page_size), self.max_page_size) if self.max_page_size else int(page_size)
            except (ValueError, TypeError):
                pass  # Fallback to default page_size if invalid
        return self.page_size

    def get_paginated_response(self, data):
        return Response({
            'count': self.page.paginator.count,
            'num_pages': self.page.paginator.num_pages,
            'page_size': self.page.paginator.per_page,
            'results': data
        })
    



class AddFarmerDataAPIViewUser(APIView):
    def post(self, request):
        print("=" * 60)
        print("🚀 ADD FARMER DATA APIVIEW - POST REQUEST STARTED")
        print("=" * 60)
        print(f"📋 Request method: {request.method}")
        print(f"📋 Request content type: {request.content_type}")
        print(f"📋 Request POST data keys: {list(request.POST.keys())}")
        print(f"📋 Request FILES keys: {list(request.FILES.keys())}")
        
        try:
            # Assuming 'data' is sent as a JSON string in form data
            data_str = request.POST.get('data')
            print(f"📄 Raw data string: {data_str[:100]}{'...' if len(data_str) > 100 else ''}")
            
            if not data_str:
                print("❌ ERROR: No data provided in request")
                return Response({'error': 'No data provided'}, status=status.HTTP_400_BAD_REQUEST)
            
            print("✅ Data string received successfully")
            data_list = json.loads(data_str)
            print(f"📊 Parsed data list length: {len(data_list)}")
            
            # Start database transaction
            with transaction.atomic():
                print("🔒 Database transaction STARTED")
                
                created_forms = []
                for index, data in enumerate(data_list):
                    print(f"\n🔄 Processing record {index + 1}/{len(data_list)} (Transaction Active)")
                    print(f"👤 Farmer Name: {data.get('farmerName', 'N/A')}")
                    print(f"📱 Mobile: {data.get('mobileNumber', 'N/A')}")
                    
                    # Extract location data (assuming single entry in each array)
                    # Extract location data (assuming IDs are sent)
                    state_id = data.get('state_id')
                    city_id = data.get('city_id')
                    tehsil_id = data.get('tehsil_id')
                    village_id = data.get('village_id')
                    verified_data = data.get('verifiedBy', [{}])[0]
                    
                    print(f"🏛️  State ID: {state_id}")
                    print(f"🏙️  City ID: {city_id}")
                    print(f"🏘️  Tehsil ID: {tehsil_id}")
                    print(f"🏘️  Village ID: {village_id}")
                    print(f"✅ Verified by data: {verified_data}")
                    
                    print("\n🌐 Creating Location Hierarchy...")
                    
                    # 1. Get State
                    if not state_id:
                        print("❌ ERROR: State ID is required")
                        raise ValueError('State ID is required')
                    state = State.objects.get(id=state_id)
                    print(f"   ✅ State retrieved - ID: {state.id}, Name: {state.name}")
                    
                    # 2. Get City with state ForeignKey
                    if not city_id:
                        print("❌ ERROR: City ID is required")
                        raise ValueError('City ID is required')
                    city = City.objects.get(id=city_id, state=state)
                    print(f"   ✅ City retrieved - ID: {city.id}, Name: {city.name}")
                    
                    # 3. Get Tehsil with city ForeignKey
                    if not tehsil_id:
                        print("❌ ERROR: Tehsil ID is required")
                        raise ValueError('Tehsil ID is required')
                    tehsil = Tehsil.objects.get(id=tehsil_id, city=city)
                    print(f"   ✅ Tehsil retrieved - ID: {tehsil.id}, Name: {tehsil.name}")
                    
                    # 4. Get Village with tehsil ForeignKey
                    if not village_id:
                        print("❌ ERROR: Village ID is required")
                        raise ValueError('Village ID is required')
                    village = Village.objects.get(id=village_id, tehsil=tehsil)
                    print(f"   ✅ Village retrieved - ID: {village.id}, Name: {village.display_name}")
                    
                    print("\n📸 Processing Images...")


                    # 5. Create VerifiedBy
                    verification_image = request.FILES.get('verification_image')
                    print(f"   🖼️  Verification image: {'✅ FOUND' if verification_image else '❌ MISSING'}")
                    
                    if not verification_image:
                        print("❌ ERROR: verification_image is required")
                        raise ValueError('verification_image is required')
                    
                    print(f"   👤 Surveyor: {verified_data.get('surveyorName')}")
                    print(f"   📅 Verification date: {verified_data.get('date')}")
                    
                    parsed_date = parse_date(verified_data.get('date'))
                    if not parsed_date:
                        print(f"❌ ERROR: Invalid date format: {verified_data.get('date')}")
                        raise ValueError(f'Invalid date format: {verified_data.get("date")}')
                    
                    verified_by = VerifiedBy.objects.create(
                        verification_image=verification_image,
                        surveyor_name=verified_data.get('surveyorName'),
                        date=parsed_date,
                        surveyor_mobile='',  # Empty string since it's optional
                        is_verified=False  # Default value
                    )
                    print(f"   ✅ VerifiedBy CREATED - ID: {verified_by.id}")
                    
                    print("\n👨‍🌾 Creating FarmerForm...")
                    
                    # 6. Create FarmerForm
                    farmer_image = request.FILES.get('farmer_image')
                    print(f"   🖼️  Farmer image: {'✅ FOUND' if farmer_image else '❌ MISSING'}")
                    
                    if not farmer_image:
                        print("❌ ERROR: farmer_image is required")
                        raise ValueError('farmer_image is required')
                    
                    # Create farmer form with all fields
                    farmer_form = FarmerForm.objects.create(
                        verified_by=verified_by,
                        farmer_image=farmer_image,
                        farmerName=data.get('farmerName'),
                        fatherName=data.get('fatherName'),
                        mobileNumber=data.get('mobileNumber'),
                        email=data.get('email'),
                        houseStatus=data.get('houseStatus'),
                        totalLandOwned=data.get('totalLandOwned'),
                        landAffected=data.get('landAffected'),
                        cropsPlanted=data.get('cropsPlanted'),
                        cropsLost=data.get('cropsLost'),
                        estimatedCropLoss=data.get('estimatedCropLoss'),
                        tractorLeveling=data.get('tractorLeveling'),
                        manureFertilizer=data.get('manureFertilizer'),
                        seedsRequired=data.get('seedsRequired'),
                        fertilizersPesticides=data.get('fertilizersPesticides'),
                        laborRequirement=data.get('laborRequirement'),
                        irrigationRepair=data.get('irrigationRepair'),
                        livestockDamage=data.get('livestockDamage'),
                        householdNeeds=data.get('householdNeeds'),
                        housingRepair=data.get('housingRepair'),
                        otherSupport=data.get('otherSupport'),
                        additionalNotes=data.get('additionalNotes'),
                        amount_needed=data.get('amount_needed')
                    )
                    
                    print(f"   ✅ FarmerForm CREATED - ID: {farmer_form.id}")
                    print(f"   👤 Farmer: {farmer_form.farmerName}")
                    print(f"   📱 Mobile: {farmer_form.mobileNumber}")
                    print(f"   🏠 House Status: {farmer_form.houseStatus}")
                    print(f"   🌾 Total Land: {farmer_form.totalLandOwned}")
                    
                    print("\n🔗 Creating TempPersonDataForm...")
                    
                    # 7. Create TempPersonDataForm
                    temp_person_data = TempPersonDataForm.objects.create(
                        village=village,
                        temp_form=farmer_form,
                        is_processed=False  # Default value
                    )
                    
                    print(f"   ✅ TempPersonDataForm CREATED - ID: {temp_person_data.id}")
                    print(f"   🏘️  Village: {village.display_name}")
                    
                    # Add to response
                    created_forms.append({
                        'id': farmer_form.id,
                        'farmerName': farmer_form.farmerName,
                        'mobileNumber': farmer_form.mobileNumber,
                        'village': village.display_name,
                        'created_at': farmer_form.created_at.isoformat()
                    })
                    
                    print(f"✅ Record {index + 1} processed successfully!")
                    print("-" * 40)
                
                print(f"\n🎉 All {len(data_list)} records processed successfully!")
                print(f"📋 Created forms count: {len(created_forms)}")
                
                print("🔓 Database transaction COMMITTED")
                
                response_data = {
                    'message': 'Data created successfully', 
                    'created_forms': created_forms,
                    'total_processed': len(created_forms),
                    'transaction_status': 'committed'
                }
                
                print("📤 Sending success response...")
                return Response(response_data, status=status.HTTP_201_CREATED)
            
        except json.JSONDecodeError as e:
            print(f"❌ JSON DECODE ERROR: {str(e)}")
            print(f"📄 Problematic data: {data_str[:200] if 'data_str' in locals() else 'N/A'}")
            return Response({'error': f'Invalid JSON in data field: {str(e)}'}, status=status.HTTP_400_BAD_REQUEST)
        
        except ValueError as e:
            print(f"❌ VALUE ERROR: {str(e)}")
            print("🔓 Database transaction ROLLEDBACK due to validation error")
            return Response({'error': f'Validation error: {str(e)}'}, status=status.HTTP_400_BAD_REQUEST)
        
        except Exception as e:
            print(f"❌ GENERAL EXCEPTION: {str(e)}")
            print("🔓 Database transaction ROLLEDBACK due to exception")
            import traceback
            print("📋 Full traceback:")
            traceback.print_exc()
            return Response({'error': f'Server error: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
        finally:
            print("=" * 60)
            print("🏁 ADD FARMER DATA APIVIEW - POST REQUEST COMPLETED")
            print("=" * 60)




class AddFarmerDataAPIViewAdmin(APIView):
    def post(self, request):
        print("=" * 60)
        print("🚀 ADD FARMER DATA ADMIN APIVIEW - POST REQUEST STARTED")
        print("=" * 60)
        print(f"📋 Request method: {request.method}")
        print(f"📋 Request content type: {request.content_type}")
        print(f"📋 Request POST data keys: {list(request.POST.keys())}")
        print(f"📋 Request FILES keys: {list(request.FILES.keys())}")
        
        try:
            # Assuming 'data' is sent as a JSON string in form data
            data_str = request.POST.get('data')
            print(f"📄 Raw data string: {data_str[:100]}{'...' if len(data_str) > 100 else ''}")
            
            if not data_str:
                print("❌ ERROR: No data provided in request")
                return Response({'error': 'No data provided'}, status=status.HTTP_400_BAD_REQUEST)
            
            print("✅ Data string received successfully")
            data_list = json.loads(data_str)
            print(f"📊 Parsed data list length: {len(data_list)}")
            
            # Start database transaction
            with transaction.atomic():
                print("🔒 Database transaction STARTED")
                
                created_forms = []
                for index, data in enumerate(data_list):
                    print(f"\n🔄 Processing record {index + 1}/{len(data_list)} (Transaction Active)")
                    print(f"👤 Farmer Name: {data.get('farmerName', 'N/A')}")
                    print(f"📱 Mobile: {data.get('mobileNumber', 'N/A')}")
                    
                    # Extract location data (assuming single entry in each array)
                    state_id = data.get('state_id')
                    city_id = data.get('city_id')
                    tehsil_id = data.get('tehsil_id')
                    village_id = data.get('village_id')
                    verified_data = data.get('verifiedBy', [{}])[0]
                    
                    print(f"🏛️  State ID: {state_id}")
                    print(f"🏙️  City ID: {city_id}")
                    print(f"🏘️  Tehsil ID: {tehsil_id}")
                    print(f"🏘️  Village ID: {village_id}")
                    print(f"✅ Verified by data: {verified_data}")
                    
                    print("\n🌐 Creating Location Hierarchy...")
                    
                    # 1. Get State
                    if not state_id:
                        print("❌ ERROR: State ID is required")
                        raise ValueError('State ID is required')
                    state = State.objects.get(id=state_id)
                    print(f"   ✅ State retrieved - ID: {state.id}, Name: {state.name}")
                    
                    # 2. Get City with state ForeignKey
                    if not city_id:
                        print("❌ ERROR: City ID is required")
                        raise ValueError('City ID is required')
                    city = City.objects.get(id=city_id, state=state)
                    print(f"   ✅ City retrieved - ID: {city.id}, Name: {city.name}")
                    
                    # 3. Get Tehsil with city ForeignKey
                    if not tehsil_id:
                        print("❌ ERROR: Tehsil ID is required")
                        raise ValueError('Tehsil ID is required')
                    tehsil = Tehsil.objects.get(id=tehsil_id, city=city)
                    print(f"   ✅ Tehsil retrieved - ID: {tehsil.id}, Name: {tehsil.name}")
                    
                    # 4. Get Village with tehsil ForeignKey
                    if not village_id:
                        print("❌ ERROR: Village ID is required")
                        raise ValueError('Village ID is required')
                    village = Village.objects.get(id=village_id, tehsil=tehsil)
                    print(f"   ✅ Village retrieved - ID: {village.id}, Name: {village.display_name}")
                    
                    print("\n📸 Processing Images...")
                    
                    # 5. Create VerifiedBy
                    verification_image = request.FILES.get('verification_image')
                    print(f"   🖼️  Verification image: {'✅ FOUND' if verification_image else '❌ MISSING'}")
                    
                    if not verification_image:
                        print("❌ ERROR: verification_image is required")
                        raise ValueError('verification_image is required')
                    
                    print(f"   👤 Surveyor: {verified_data.get('surveyorName')}")
                    print(f"   📅 Verification date: {verified_data.get('date')}")
                    
                    parsed_date = parse_date(verified_data.get('date'))
                    if not parsed_date:
                        print(f"❌ ERROR: Invalid date format: {verified_data.get('date')}")
                        raise ValueError(f'Invalid date format: {verified_data.get("date")}')
                    
                    verified_by = VerifiedBy.objects.create(
                        verification_image=verification_image if verification_image else None,
                        surveyor_name=verified_data.get('surveyorName'),
                        date=parsed_date,
                        surveyor_mobile='',  # Empty string since it's optional
                        is_verified=True  # Will be set to True in processing step
                    )
                    print(f"   ✅ VerifiedBy CREATED - ID: {verified_by.id}")
                    
                    print("\n👨‍🌾 Creating FarmerForm...")
                    
                    # 6. Create FarmerForm
                    farmer_image = request.FILES.get('farmer_image')
                    print(f"   🖼️  Farmer image: {'✅ FOUND' if farmer_image else '❌ MISSING'}")
                    
                    if not farmer_image:
                        print("❌ ERROR: farmer_image is required")
                        raise ValueError('farmer_image is required')
                    
                    # Create farmer form with all fields
                    farmer_form = FarmerForm.objects.create(
                        verified_by=verified_by,
                        farmer_image=farmer_image,
                        farmerName=data.get('farmerName'),
                        fatherName=data.get('fatherName'),
                        mobileNumber=data.get('mobileNumber'),
                        email=data.get('email'),
                        houseStatus=data.get('houseStatus'),
                        totalLandOwned=data.get('totalLandOwned'),
                        landAffected=data.get('landAffected'),
                        cropsPlanted=data.get('cropsPlanted'),
                        cropsLost=data.get('cropsLost'),
                        estimatedCropLoss=data.get('estimatedCropLoss'),
                        tractorLeveling=data.get('tractorLeveling'),
                        manureFertilizer=data.get('manureFertilizer'),
                        seedsRequired=data.get('seedsRequired'),
                        fertilizersPesticides=data.get('fertilizersPesticides'),
                        laborRequirement=data.get('laborRequirement'),
                        irrigationRepair=data.get('irrigationRepair'),
                        livestockDamage=data.get('livestockDamage'),
                        householdNeeds=data.get('householdNeeds'),
                        housingRepair=data.get('housingRepair'),
                        otherSupport=data.get('otherSupport'),
                        additionalNotes=data.get('additionalNotes'),
                        amount_needed = data.get('amount_needed'),
                        is_active=True
                    )
                    
                    print(f"   ✅ FarmerForm CREATED - ID: {farmer_form.id}")
                    print(f"   👤 Farmer: {farmer_form.farmerName}")
                    print(f"   📱 Mobile: {farmer_form.mobileNumber}")
                    print(f"   🏠 House Status: {farmer_form.houseStatus}")
                    print(f"   🌾 Total Land: {farmer_form.totalLandOwned}")
                    
                    print("\n👤 Creating User Account...")
                    # 7. Create User
                    email = farmer_form.email if farmer_form.email else ''
                    username = email if email else f"{farmer_form.farmerName.replace(' ', '_').lower()}"
                    
                    # Ensure unique username
                    original_username = username
                    counter = 1
                    while User.objects.filter(username=username).exists():
                        username = f"{original_username}_{counter}"
                        counter += 1
                    
                    print(f"   📧 Email: {email}")
                    print(f"   👤 Username: {username}")
                    
                    user = User.objects.create_user(
                        username=username,
                        email=email if email else '',
                        password='abc123',  # Temporary password - should be changed
                        first_name=farmer_form.farmerName.split()[0] if farmer_form.farmerName else '',
                        last_name=' '.join(farmer_form.farmerName.split()[1:]) if farmer_form.farmerName and len(farmer_form.farmerName.split()) > 1 else ''
                    )
                    print(f"   ✅ User CREATED - ID: {user.id}, Username: {user.username}")
                    
                    print("\n📝 Creating PersonData...")
                    
                    # 8. Create PersonData (replaces TempPersonDataForm)
                    person_data = PersonData.objects.create(
                        village=village,
                        form=farmer_form,
                        user=user,
                        is_active=True,
                        is_single_user=True  # As this is admin creation
                    )
                    
                    print(f"   ✅ PersonData CREATED - ID: {person_data.id}")
                    print(f"   ✅ is_single_user: {person_data.is_single_user}")
                    print(f"   🏘️  Village: {village.display_name}")
                    
                    # Add to response
                    created_forms.append({
                        'id': farmer_form.id,
                        'farmerName': farmer_form.farmerName,
                        'mobileNumber': farmer_form.mobileNumber,
                        'village': village.display_name,
                        'user': {
                            'id': user.id,
                            'username': user.username,
                            'email': user.email
                        },
                        'person_data_id': person_data.id,
                        'created_at': farmer_form.created_at.isoformat()
                    })
                    
                    print(f"✅ Record {index + 1} processed successfully!")
                    print("-" * 40)
                
                print(f"\n🎉 All {len(data_list)} records processed successfully!")
                print(f"📋 Created forms count: {len(created_forms)}")
                
                print("🔓 Database transaction COMMITTED")
                
                response_data = {
                    'message': 'Data created successfully', 
                    'created_forms': created_forms,
                    'total_processed': len(created_forms),
                    'transaction_status': 'committed',
                    'users_created': len(created_forms),
                    'person_data_created': len(created_forms)
                }
                
                print("📤 Sending success response...")
                return Response(response_data, status=status.HTTP_201_CREATED)
            
        except json.JSONDecodeError as e:
            print(f"❌ JSON DECODE ERROR: {str(e)}")
            print(f"📄 Problematic data: {data_str[:200] if 'data_str' in locals() else 'N/A'}")
            return Response({'error': f'Invalid JSON in data field: {str(e)}'}, status=status.HTTP_400_BAD_REQUEST)
        
        except ValueError as e:
            print(f"❌ VALUE ERROR: {str(e)}")
            print("🔓 Database transaction ROLLEDBACK due to validation error")
            return Response({'error': f'Validation error: {str(e)}'}, status=status.HTTP_400_BAD_REQUEST)
        
        except Exception as e:
            print(f"❌ GENERAL EXCEPTION: {str(e)}")
            print("🔓 Database transaction ROLLEDBACK due to exception")
            import traceback
            print("📋 Full traceback:")
            traceback.print_exc()
            return Response({'error': f'Server error: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
        finally:
            print("=" * 60)
            print("🏁 ADD FARMER DATA ADMIN APIVIEW - POST REQUEST COMPLETED")
            print("=" * 60)




class ProcessTempPersonDataAPIView(APIView):
    def get(self, request):
        """
        GET: Retrieve processed PersonData records with details
        """
        print("=" * 60)
        print("🔄 PROCESS TEMP PERSON DATA - GET REQUEST")
        print("=" * 60)
        
        try:
            # Get all processed TempPersonDataForms with related data
            temp_forms = TempPersonDataForm.objects.filter(
                is_processed=True
            ).select_related(
                'village', 'temp_form', 'temp_form__verified_by'
            ).prefetch_related(
                'temp_person__form', 'temp_person__village', 'temp_person__user'
            )
            
            processed_data = []
            for temp_form in temp_forms:
                # Get related PersonData if exists
                person_data = temp_form.temp_person.first()
                
                # Get location hierarchy
                village = temp_form.village
                tehsil = village.tehsil if village else None
                city = tehsil.city if tehsil else None
                state = city.state if city else None
                
                # Build location data
                village_data = {
                    "name": village.display_name if village else "",
                    "pincode": village.pin_code if village else "",
                    "latitude": village.latitude if village else "",
                    "longitude": village.longitude if village else ""
                } if village else {}
                
                tehsil_data = {"name": tehsil.name} if tehsil else {}
                city_data = {"name": city.name} if city else {}
                state_data = {"name": state.name} if state else {}
                
                # Get verified by data
                verified_by = temp_form.temp_form.verified_by if temp_form.temp_form else None
                verified_data = {}
                if verified_by:
                    verified_data = {
                        "surveyorName": verified_by.surveyor_name,
                        "date": verified_by.date.strftime('%Y-%m-%d') if verified_by.date else "",
                        "is_verified": verified_by.is_verified
                    }
                
                # Get farmer form data
                farmer_form = temp_form.temp_form
                form_data = {}
                if farmer_form:
                    form_data = {
                        "farmerName": farmer_form.farmerName,
                        "fatherName": farmer_form.fatherName,
                        "mobileNumber": farmer_form.mobileNumber,
                        "email": farmer_form.email,
                        "houseStatus": farmer_form.houseStatus,
                        "totalLandOwned": farmer_form.totalLandOwned,
                        "landAffected": farmer_form.landAffected,
                        "cropsPlanted": farmer_form.cropsPlanted,
                        "cropsLost": farmer_form.cropsLost,
                        "estimatedCropLoss": farmer_form.estimatedCropLoss,
                        "tractorLeveling": farmer_form.tractorLeveling,
                        "manureFertilizer": farmer_form.manureFertilizer,
                        "seedsRequired": farmer_form.seedsRequired,
                        "fertilizersPesticides": farmer_form.fertilizersPesticides,
                        "laborRequirement": farmer_form.laborRequirement,
                        "irrigationRepair": farmer_form.irrigationRepair,
                        "livestockDamage": farmer_form.livestockDamage,
                        "householdNeeds": farmer_form.householdNeeds,
                        "housingRepair": farmer_form.housingRepair,
                        "otherSupport": farmer_form.otherSupport,
                        "additionalNotes": farmer_form.additionalNotes,
                        "is_active": farmer_form.is_active
                    }
                
                # Get user data if exists
                user_data = {}
                if person_data and person_data.user:
                    user_data = {
                        "id": person_data.user.id,
                        "username": person_data.user.username,
                        "email": person_data.user.email,
                        "first_name": person_data.user.first_name,
                        "last_name": person_data.user.last_name,
                        "date_joined": person_data.user.date_joined.isoformat() if person_data.user.date_joined else ""
                    }
                
                record_data = {
                    "temp_form_id": temp_form.id,
                    "processed": temp_form.is_processed,
                    "village": [village_data],
                    "state": [state_data],
                    "city": [city_data],
                    "tehsil": [tehsil_data],
                    "farmer_form": form_data,
                    "verifiedBy": [verified_data],
                    "person_data": {
                        "id": person_data.id if person_data else None,
                        "is_active": person_data.is_active if person_data else True,
                        "is_single_user": person_data.is_single_user if person_data else False,
                        "user": user_data
                    },
                    "created_at": temp_form.created_at.isoformat(),
                    "updated_at": temp_form.updated_at.isoformat()
                }
                
                processed_data.append(record_data)
            
            print(f"📊 Retrieved {len(processed_data)} processed records")
            print("📤 Sending GET response...")
            
            return Response({
                'message': 'Processed person data retrieved successfully',
                'data': processed_data,
                'total_records': len(processed_data)
            }, status=status.HTTP_200_OK)
        
        except Exception as e:
            print(f"❌ GET REQUEST ERROR: {str(e)}")
            print("📋 Full traceback:")
            traceback.print_exc()
            return Response({
                'error': f'Server error: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
        finally:
            print("=" * 60)
            print("🏁 PROCESS TEMP PERSON DATA - GET REQUEST COMPLETED")
            print("=" * 60)

    def post(self, request):
        """
        POST: Process TempPersonDataForm by ID - Create PersonData, User, and update status
        """
        print("=" * 60)
        print("🚀 PROCESS TEMP PERSON DATA - POST REQUEST STARTED")
        print("=" * 60)
        print(f"📋 Request method: {request.method}")
        print(f"📋 Request data: {request.data}")
        
        try:
            # Get temp_person_data_form ID from request
            temp_form_id = request.data.get('temp_person_data_form_id')
            if not temp_form_id:
                print("❌ ERROR: temp_person_data_form_id is required")
                return Response({
                    'error': 'temp_person_data_form_id is required'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            print(f"🔍 Looking for TempPersonDataForm ID: {temp_form_id}")
            
            # Get the TempPersonDataForm instance
            try:
                temp_form = TempPersonDataForm.objects.select_related(
                    'village', 'temp_form', 'temp_form__verified_by'
                ).get(id=temp_form_id)
            except TempPersonDataForm.DoesNotExist:
                print(f"❌ ERROR: TempPersonDataForm with ID {temp_form_id} not found")
                return Response({
                    'error': f'TempPersonDataForm with ID {temp_form_id} not found'
                }, status=status.HTTP_404_NOT_FOUND)
            
            print(f"✅ Found TempPersonDataForm - ID: {temp_form.id}")
            print(f"👤 Farmer: {temp_form.temp_form.farmerName}")
            print(f"📱 Mobile: {temp_form.temp_form.mobileNumber}")
            print(f"🏘️  Village: {temp_form.village.display_name}")
            print(f"📋 Current status - Processed: {temp_form.is_processed}")
            
            # Check if already processed
            if temp_form.is_processed:
                print("⚠️  WARNING: This form is already processed")
                return Response({
                    'message': 'This form has already been processed',
                    'temp_form_id': temp_form.id,
                    'status': 'already_processed'
                }, status=status.HTTP_200_OK)
            
            # Start database transaction for atomic operations
            with transaction.atomic():
                print("🔒 Database transaction STARTED")
                
                # 1. Update VerifiedBy - Set is_verified = True
                print("✅ Updating VerifiedBy status...")
                verified_by = temp_form.temp_form.verified_by
                verified_by.is_verified = True
                verified_by.save()
                print(f"   ✅ VerifiedBy updated - ID: {verified_by.id}, is_verified: {verified_by.is_verified}")
                
                # 2. Create User
                print("👤 Creating User...")
                farmer_form = temp_form.temp_form
                email = farmer_form.email if farmer_form.email else ''
                username = email if email else f"{farmer_form.farmerName.replace(' ', '_').lower()}"
                
                # Ensure unique username
                original_username = username
                counter = 1
                while User.objects.filter(username=username).exists():
                    username = f"{original_username}_{counter}"
                    counter += 1
                
                print(f"   📧 Email: {email}")
                print(f"   👤 Username: {username}")
                
                user = User.objects.create_user(
                    username=username,
                    email=email if email else '',
                    password='abc123',  # Temporary password - should be changed
                    first_name=farmer_form.farmerName.split()[0] if farmer_form.farmerName else '',
                    last_name=' '.join(farmer_form.farmerName.split()[1:]) if farmer_form.farmerName and len(farmer_form.farmerName.split()) > 1 else '',
                    is_person=True
                )
                print(f"   ✅ User CREATED - ID: {user.id}, Username: {user.username}")
                
                # 3. Create PersonData
                print("📝 Creating PersonData...")
                person_data = PersonData.objects.create(
                    temp_person_data_form=temp_form,
                    village=temp_form.village,
                    form=farmer_form,
                    user=user,
                    is_active=True,
                    is_single_user=True  # As requested
                )
                print(f"   ✅ PersonData CREATED - ID: {person_data.id}")
                print(f"   ✅ is_single_user: {person_data.is_single_user}")
                
                # 4. Update TempPersonDataForm - Set is_processed = True
                print("📋 Updating TempPersonDataForm status...")
                temp_form.is_processed = True
                temp_form.save()
                print(f"   ✅ TempPersonDataForm updated - ID: {temp_form.id}, is_processed: {temp_form.is_processed}")
                
                print("🔓 Database transaction COMMITTED")
                
                # Build location hierarchy for response
                village = temp_form.village
                tehsil = village.tehsil if village else None
                city = tehsil.city if tehsil else None
                state = city.state if city else None
                
                village_data = {
                    "name": village.display_name,
                    "pincode": village.pin_code,
                    "latitude": village.latitude,
                    "longitude": village.longitude
                }
                
                tehsil_data = {"name": tehsil.name} if tehsil else {}
                city_data = {"name": city.name} if city else {}
                state_data = {"name": state.name} if state else {}
                
                # Build response
                response_data = {
                    'message': 'TempPersonDataForm processed successfully',
                    'temp_form_id': temp_form.id,
                    'person_data': {
                        'id': person_data.id,
                        'is_active': person_data.is_active,
                        'is_single_user': person_data.is_single_user,
                        'user': {
                            'id': user.id,
                            'username': user.username,
                            'email': user.email,
                            'first_name': user.first_name,
                            'last_name': user.last_name,
                            'date_joined': user.date_joined.isoformat()
                        }
                    },
                    'updated_records': {
                        'temp_person_data_form': {
                            'id': temp_form.id,
                            'is_processed': temp_form.is_processed
                        },
                        'verified_by': {
                            'id': verified_by.id,
                            'is_verified': verified_by.is_verified
                        }
                    },
                    'farmer_info': {
                        'farmerName': farmer_form.farmerName,
                        'mobileNumber': farmer_form.mobileNumber,
                        'email': farmer_form.email,
                        'village': [village_data],
                        'state': [state_data],
                        'city': [city_data],
                        'tehsil': [tehsil_data]
                    },
                    'transaction_status': 'committed'
                }
                
                print("✅ Processing completed successfully!")
                print("📤 Sending success response...")
                
                return Response(response_data, status=status.HTTP_201_CREATED)
        
        except ValueError as e:
            print(f"❌ VALIDATION ERROR: {str(e)}")
            print("🔓 Database transaction ROLLEDBACK due to validation error")
            return Response({
                'error': f'Validation error: {str(e)}'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        except Exception as e:
            print(f"❌ GENERAL EXCEPTION: {str(e)}")
            print("🔓 Database transaction ROLLEDBACK due to exception")
            print("📋 Full traceback:")
            traceback.print_exc()
            return Response({
                'error': f'Server error: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
        finally:
            print("=" * 60)
            print("🏁 PROCESS TEMP PERSON DATA - POST REQUEST COMPLETED")
            print("=" * 60) 






class ShowPersonTemporaryAll(APIView):
    pagination_class = TempPersonPagination
    filter_backends = [DjangoFilterBackend]
    filterset_class = TempPersonDataFormFilter

    def get(self, request):
        # Get all TempPersonDataForm instances
        temp_person = TempPersonDataForm.objects.all()

        # Apply filters
        filterset = self.filterset_class(request.query_params, queryset=temp_person)
        if not filterset.is_valid():
            return Response(filterset.errors, status=400)
        temp_person = filterset.qs

        # Apply pagination
        paginator = self.pagination_class()
        page = paginator.paginate_queryset(temp_person, request)

        data = []

        for person_data in page:
            data.append({
                'temp_person_id': person_data.id,
                'is_processed': person_data.is_processed,
                "village_data": [{
                    'village_id': person_data.village.id,
                    'display_name': person_data.village.display_name,
                    'pin_code': person_data.village.pin_code
                }],
                "form_data": [{
                    "form_id": person_data.temp_form.id,
                    'farmer_name': person_data.temp_form.farmerName,
                    'farmer_image': urljoin(
                        request.build_absolute_uri('/'),
                        person_data.temp_form.farmer_image.url
                    ),
                    'mobileNumber': person_data.temp_form.mobileNumber,
                }]
            })

        return paginator.get_paginated_response(data)



class ShowPersonTemporaryById(APIView):
    def get(self,request):
        temp_person_id = request.query_params.get('temp_person_id')

        temp_person = TempPersonDataForm.objects.filter(id=temp_person_id)

        
        data = []

        for person_data in temp_person:
            data.append({
                'temp_person_id':person_data.id,
                'is_processed':person_data.is_processed,
                "village_data":[{
                    'village_id':person_data.village.id,
                    'display_name':person_data.village.display_name,
                    'pin_code':person_data.village.pin_code,
                    'longitude':person_data.village.longitude,
                    'latitude':person_data.village.latitude,
                    "tehsil_data":[{
                        'tehsil_id':person_data.village.tehsil.id,
                        'name':person_data.village.tehsil.name,
                        'city_data':[{
                            'city_id':person_data.village.tehsil.city.id,
                            'name':person_data.village.tehsil.city.name,
                            'state_data':[{
                                'state_id':person_data.village.tehsil.city.state.id,
                                'name':person_data.village.tehsil.city.state.name
                            }]
                        }]
                    }]
                }],
                "form_data":[{
                    "form_id":person_data.temp_form.id,
                    'farmer_name':person_data.temp_form.farmerName,
                    'farmer_image':urljoin(
                    request.build_absolute_uri('/'), 
                    person_data.temp_form.farmer_image.url
                    ),
                    'mobileNumber':person_data.temp_form.mobileNumber,
                    'fatherName':person_data.temp_form.fatherName,
                    'email':person_data.temp_form.email,
                    'houseStatus':person_data.temp_form.houseStatus,
                    'totalLandOwned':person_data.temp_form.totalLandOwned,
                    'landAffected':person_data.temp_form.landAffected,
                    'cropsPlanted':person_data.temp_form.cropsPlanted,
                    'cropsLost':person_data.temp_form.cropsLost,
                    'estimatedCropLoss':person_data.temp_form.estimatedCropLoss,
                    'tractorLeveling':person_data.temp_form.tractorLeveling,
                    'manureFertilizer':person_data.temp_form.manureFertilizer,
                    'seedsRequired':person_data.temp_form.seedsRequired,
                    'fertilizersPesticides':person_data.temp_form.fertilizersPesticides,
                    'laborRequirement':person_data.temp_form.laborRequirement,
                    'irrigationRepair':person_data.temp_form.irrigationRepair,
                    'livestockDamage':person_data.temp_form.livestockDamage,
                    'householdNeeds':person_data.temp_form.householdNeeds,
                    'housingRepair':person_data.temp_form.housingRepair,
                    'otherSupport':person_data.temp_form.otherSupport,
                    'additionalNotes':person_data.temp_form.additionalNotes,
                    'verified_by_data':[{
                        'verification_id':person_data.temp_form.verified_by.id,
                        'verification_image':urljoin(
                        request.build_absolute_uri('/'), 
                        person_data.temp_form.verified_by.verification_image.url
                        ) if person_data.temp_form.verified_by.verification_image else None,
                        'surveyor_name':person_data.temp_form.verified_by.surveyor_name,
                        'surveyor_mobile':person_data.temp_form.verified_by.surveyor_mobile,
                        'date':person_data.temp_form.verified_by.date,
                        'is_verified':person_data.temp_form.verified_by.is_verified,
                    }]
                }]
            })

        return Response(data,status=status.HTTP_200_OK)


class ShowPersonAll(APIView):
    pagination_class = TempPersonPagination

    def get(self, request):
        # For GET request, return all FarmerForm instances in the same structure as payload
        
        person = PersonData.objects.all()

        # Apply pagination
        paginator = self.pagination_class()
        page = paginator.paginate_queryset(person, request)


        data = []

        for person_data in page:
            data.append({
                'person_id':person_data.id,
                'is_active':person_data.is_active,
                'is_single_user':person_data.is_single_user,
                "village_data":[{
                    'village_id':person_data.village.id if person_data.village else None,
                    'display_name':person_data.village.display_name if person_data.village.display_name else None,
                    'pin_code':person_data.village.pin_code
                }],
                "form_data":[{
                    "form_id":person_data.form.id,
                    'farmer_name':person_data.form.farmerName,
                    'farmer_image':urljoin(
                    request.build_absolute_uri('/'), 
                    person_data.form.farmer_image.url
                    )if person_data.form and person_data.form.farmer_image else None,
                    'mobileNumber':person_data.form.mobileNumber,
                }]
            })

        return paginator.get_paginated_response(data)


class ShowPersonById(APIView):
    def get(self,request):
        person_id = request.query_params.get('person_id')

        person = PersonData.objects.filter(id=person_id)

        

        data = []

        for person_data in person:

            verification_data = None  

            if person_data.is_single_user:
                try:
                    verified_by = person_data.form.verified_by
                    verification_data = [{
                        'verified_by_data': [{
                            'verification_id': verified_by.id,
                            'verification_image': urljoin(
                                request.build_absolute_uri('/'),
                                verified_by.verification_image.url
                            ) if verified_by.verification_image else None,
                            'surveyor_name': verified_by.surveyor_name,
                            'surveyor_mobile': verified_by.surveyor_mobile,
                            'date': verified_by.date,
                            'is_verified': verified_by.is_verified,
                        }]
                    }]
                except AttributeError:
                    verification_data = None
                

            data.append({
                'person_id':person_data.id,
                'is_active':person_data.is_active,
                'is_single_user':person_data.is_single_user,
                "village_data":[{
                    'village_id':person_data.village.id,
                    'display_name':person_data.village.display_name,
                    'pin_code':person_data.village.pin_code,
                    'longitude':person_data.village.longitude,
                    'latitude':person_data.village.latitude,
                    "tehsil_data":[{
                        'tehsil_id':person_data.village.tehsil.id,
                        'name':person_data.village.tehsil.name,
                        'city_data':[{
                            'city_id':person_data.village.tehsil.city.id,
                            'name':person_data.village.tehsil.city.name,
                            'state_data':[{
                                'state_id':person_data.village.tehsil.city.state.id,
                                'name':person_data.village.tehsil.city.state.name
                            }]
                        }]
                    }]
                }],
                "form_data":[{
                    "form_id":person_data.form.id,
                    'farmer_name':person_data.form.farmerName,
                    'farmer_image':urljoin(
                    request.build_absolute_uri('/'), 
                    person_data.form.farmer_image.url
                    ) if person_data.form.farmer_image else None,
                    'mobileNumber':person_data.form.mobileNumber,
                    'fatherName':person_data.form.fatherName,
                    'email':person_data.form.email,
                    'houseStatus':person_data.form.houseStatus,
                    'totalLandOwned':person_data.form.totalLandOwned,
                    'landAffected':person_data.form.landAffected,
                    'cropsPlanted':person_data.form.cropsPlanted,
                    'cropsLost':person_data.form.cropsLost,
                    'estimatedCropLoss':person_data.form.estimatedCropLoss,
                    'tractorLeveling':person_data.form.tractorLeveling,
                    'manureFertilizer':person_data.form.manureFertilizer,
                    'seedsRequired':person_data.form.seedsRequired,
                    'fertilizersPesticides':person_data.form.fertilizersPesticides,
                    'laborRequirement':person_data.form.laborRequirement,
                    'irrigationRepair':person_data.form.irrigationRepair,
                    'livestockDamage':person_data.form.livestockDamage,
                    'householdNeeds':person_data.form.householdNeeds,
                    'housingRepair':person_data.form.housingRepair,
                    'otherSupport':person_data.form.otherSupport,
                    'additionalNotes':person_data.form.additionalNotes,
                }],
                'verification_data': verification_data 
            })

        return Response(data,status=status.HTTP_200_OK)



class TempPersonList(APIView):

    pagination_class = TempPersonPagination

    def get(self, request):
        # For GET request, return all FarmerForm instances in the same structure as payload
        
        temp_person = TempPersonDataForm.objects.all().order_by('-created_at')

        # Apply pagination
        paginator = self.pagination_class()
        page = paginator.paginate_queryset(temp_person, request)

        data = []

        for person_data in page:
            data.append({
                'temp_person_id':person_data.id,
                'is_processed':person_data.is_processed,
                "village_data":[{
                    'village_id':person_data.village.id,
                    'display_name':person_data.village.display_name,
                    'pin_code':person_data.village.pin_code
                }],
                "form_data":[{
                    "form_id":person_data.temp_form.id,
                    'farmer_name':person_data.temp_form.farmerName,
                    'farmer_image':urljoin(
                    request.build_absolute_uri('/'), 
                    person_data.temp_form.farmer_image.url
                    ) if person_data.temp_form.farmer_image else None,
                    'mobileNumber':person_data.temp_form.mobileNumber,
                }]
            })

        return paginator.get_paginated_response(data)




class UniquePinCodesView(APIView):
    def get(self, request):
        try:
            # Query distinct pin codes and their associated village IDs
            pin_codes_with_ids = Village.objects.values('pin_code').distinct()
            result = []
            for entry in pin_codes_with_ids:
                pin_code = entry['pin_code']
                village_ids = Village.objects.filter(pin_code=pin_code).values_list('id', flat=True)
                result.append({
                    'pin_code': pin_code,
                    'village_ids': list(village_ids)
                })
            return Response({'pin_codes':   result}, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class PersonDataByPinCodeView(APIView):
    def get(self, request):
        pin_code = request.query_params.get('pin_code')
        print("\n==== Incoming GET Request ====")
        print("Query Params:", request.query_params)
        print("Pin Code received:", pin_code)

        if not pin_code:
            print("❌ No pin_code provided in request.")
            return Response({'error': 'pin_code is required'}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            # Filter PersonData by village pin_code
            person_data_queryset = PersonData.objects.filter(village__pin_code=pin_code)
            print("\n✅ person_data_queryset created")
            print("QuerySet (SQL):", str(person_data_queryset.query))
            print("Count before fetching:", person_data_queryset.count())

            # Total count of PersonData entries for this pin_code
            total_person_data_count = person_data_queryset.count()
            print("📊 Total PersonData Count:", total_person_data_count)
            
            # Get all FarmerForms (including duplicates)
            farmers_data = person_data_queryset.values(
                'form__id', 'form__farmerName'
            )
            print("\n📜 Farmers Data (Raw):", list(farmers_data)[:5], "...")  # Show first 5 for preview

            # Fetch unique location data
            location_data = person_data_queryset.values('village__longitude', 'village__latitude').distinct()
            print("\n🌍 Location Data (Raw):", list(location_data))

            location_list = [
                {
                    'longitude': location['village__longitude'],
                    'latitude': location['village__latitude']
                } for location in location_data
            ]
            print("📍 Parsed Location List:", location_list)
            
            # Total count of all FarmerForm entries for this pin_code
            total_farmers_count = farmers_data.count()
            print("👨‍🌾 Total Farmers Count:", total_farmers_count)
            
            # Prepare the list of farmers
            farmers_list = [
                {
                    'form_id': entry['form__id'],
                    'farmer_name': entry['form__farmerName']
                } for entry in farmers_data
            ]
            print("✅ Farmers List (Processed):", farmers_list[:5], "...")  # Preview first 5
            
            response_data = {
                'pin_code': pin_code,
                'total_person_data_count': total_person_data_count,
                'total_farmers_count': total_farmers_count,
                'farmers': farmers_list,
                'location_list': location_list
            }
            
            print("\n📦 Final Response Data:", response_data)
            return Response(response_data, status=status.HTTP_200_OK)

        except Exception as e:
            print("❌ Exception occurred while processing:", str(e))
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class VillageDisplayNamesByPinCodeView(APIView):
    def get(self, request):
        pin_code = request.query_params.get('pin_code')
        search_query = request.query_params.get('search', '')  # Get search parameter, default to empty string
        
        if not pin_code:
            return Response({'error': 'pin_code is required'}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            # Query distinct display names for the given pin_code
            queryset = Village.objects.filter(pin_code=pin_code).values('display_name').distinct()
            
            # Apply search filter if search_query is provided
            if search_query:
                queryset = queryset.filter(display_name__icontains=search_query)
            
            # Extract display_name values into a list
            display_names = [entry['display_name'] for entry in queryset]
            
            response_data = {
                'pin_code': pin_code,
                'display_names': display_names
            }
            
            return Response(response_data, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class AddAffectedVillageMapData(APIView):
    def post(self, request):
        pin_code = request.data.get('pin_code','')
        popup = request.data.get('popup')
        center = request.data.get('center')
        zoom = request.data.get('zoom')
        marker = request.data.get('marker')
        radius = request.data.get('radius')
        severity = request.data.get('severity')
        population = request.data.get('population')
        village_id = request.data.get('village_id')

        try:
            village = Village.objects.get(id=village_id)
        except Village.DoesNotExist:
            return Response("Village Not Found",status=status.HTTP_404_NOT_FOUND)

        required_fields = ['popup', 'center', 'marker', 'severity', 'population','village_id']
        for field in required_fields:
            if not request.data.get(field):
                return Response(
                    {"error": f"{field} is required"},
                    status=status.HTTP_400_BAD_REQUEST
                )

        try:
            AffectedVillageMapData.objects.get(village=village)
            return Response(
                {"error": "This Village Data already exists ... Please Update it"},
                status=status.HTTP_400_BAD_REQUEST
            )
        except AffectedVillageMapData.DoesNotExist:
            try:
                affected_data_creation = AffectedVillageMapData.objects.create(
                    pin_code=pin_code or '',
                    popup=popup,
                    center=center,
                    zoom=zoom or '14',  
                    marker=marker,
                    radius=radius or '600', 
                    severity=severity,
                    population=population,
                    village=village
                )

                response_data = {
                    'pin_code': affected_data_creation.pin_code,
                    'popup': affected_data_creation.popup,
                    'center': affected_data_creation.center,
                    'zoom': affected_data_creation.zoom,
                    'marker': affected_data_creation.marker,
                    'radius': affected_data_creation.radius,
                    'severity': affected_data_creation.severity,
                    'population': affected_data_creation.population,
                    'village':affected_data_creation.village.display_name,
                    'village_id':affected_data_creation.village.id

                }

                return Response(response_data, status=status.HTTP_201_CREATED)
            except Exception as e:
                return Response(
                    {"error": f"Failed to create record: {str(e)}"},
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR
                )
        except AffectedVillageMapData.MultipleObjectsReturned:
            return Response(
                {"error": "Multiple records found for this pin_code. Please ensure pin_code is unique"},
                status=status.HTTP_400_BAD_REQUEST
            )
        except Exception as e:
            return Response(
                {"error": f"An unexpected error occurred: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        


class AffectedVillageMapDataView(APIView):
    def get(self, request):
        pin_code = request.query_params.get('pin_code')
        popup = request.query_params.get('popup')

       

        try:
            # Initialize queryset
            queryset = AffectedVillageMapData.objects.all()

            # Filter by pin_code or popup if provided
            if pin_code:
                queryset = queryset.filter(pin_code=pin_code)
            if popup:
                queryset = queryset.filter(popup=popup)

            
             # Check if queryset is empty
            if not queryset.exists():
                # Construct dummy response
                dummy_response = {
                    "VARPAL KHURD": {
                        "center": {"latitude": 31.6008513, "longitude": 74.6060595},
                        "zoom": 14,
                        "marker": {"latitude": 31.6008513, "longitude": 74.6060595},
                        "popup": "VARPAL KHURD",
                        "radius": 600,
                        "severity": "medium",
                        "population": 10
                    }
                }
                return Response(dummy_response, status=status.HTTP_200_OK)


            # Prepare response as a dictionary with popup as keys
            response_data = {}
            for record in queryset:
                # Use center and marker directly as dictionaries
                center = record.center if record.center else {}
                marker = record.marker if record.marker else {}

                # Ensure center and marker are dictionaries
                if not isinstance(center, dict) or not isinstance(marker, dict):
                    return Response(
                        {"No Response Found"},
                        status=status.HTTP_200_OK
                    )

                response_data[record.popup] = {
                    'center': center,
                    'zoom': int(record.zoom) ,
                    'marker': marker,
                    'popup': record.popup ,
                    'radius': int(record.radius) if record.radius else 600,
                    'severity': record.severity ,
                    'population': int(record.population) if record.population else 0
                }

            

            return Response(response_data, status=status.HTTP_200_OK)
        except Exception as e:
            return Response(
                {"error": f"An unexpected error occurred: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )




class AffectedVillageMapDataViewAdmin(APIView):
    def get(self, request):
        pin_code = request.query_params.get('pin_code')
        popup = request.query_params.get('popup')

        try:
            # Initialize queryset
            queryset = AffectedVillageMapData.objects.all()

            # Filter by pin_code or popup if provided
            if pin_code:
                queryset = queryset.filter(pin_code=pin_code)
            if popup:
                queryset = queryset.filter(popup=popup)

            # Prepare response as a dictionary with popup as keys
            response_data = {}
            total_population = 0
            
            for record in queryset:
                # Use center and marker directly as dictionaries
                center = record.center if record.center else {}
                marker = record.marker if record.marker else {}

                # Ensure center and marker are dictionaries
                if not isinstance(center, dict) or not isinstance(marker, dict):
                    return Response(
                        {"error": f"Invalid format for center or marker in record with popup: {record.popup}"},
                        status=status.HTTP_400_BAD_REQUEST
                    )

                response_data[record.popup] = {
                    'center': center,
                    'zoom': int(record.zoom) if record.zoom else 14,
                    'marker': marker,
                    'popup': record.popup,
                    'radius': int(record.radius) if record.radius else 600,
                    'severity': record.severity,
                    'population': int(record.population) if record.population else 0
                }
                
                # Add to total population
                total_population += int(record.population) if record.population else 0

            if not response_data:
                return Response(
                    {"error": "No data found for the provided pin_code or popup"},
                    status=status.HTTP_404_NOT_FOUND
                )

            # Add total counts to response
            response_data['total_villages'] = len(response_data)
            response_data['total_population'] = total_population

            return Response(response_data, status=status.HTTP_200_OK)
        except Exception as e:
            return Response(
                {"error": f"An unexpected error occurred: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class UpdateAffectedVillagePopulation(APIView):
    def patch(self, request):
        village_id = request.data.get('village_id')
        population_increment = request.data.get('population_increament')

        # Validate required fields
        required_fields = ['village_id', 'population_increament']
        for field in required_fields:
            if not request.data.get(field):
                return Response(
                    {"error": f"{field} is required"},
                    status=status.HTTP_400_BAD_REQUEST
                )

        # Validate population_increment is a valid number
        try:
            population_increment = int(population_increment)
            if population_increment < 0:
                return Response(
                    {"error": "population_increment must be a non-negative number"},
                    status=status.HTTP_400_BAD_REQUEST
                )
        except (ValueError, TypeError):
            return Response(
                {"error": "population_increment must be a valid integer"},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            village = Village.objects.get(id=village_id)
        except Village.DoesNotExist:
            return Response("Village Not Found",status=status.HTTP_404_NOT_FOUND)

        try:
            # Fetch AffectedVillageMapData records for the given pin_code
            affected_village_data = AffectedVillageMapData.objects.filter(village=village)

            if not affected_village_data.exists():
                return Response(
                    {"error": f"No records found for Village: {village_id}"},
                    status=status.HTTP_404_NOT_FOUND
                )

            updated_records = []
            for record in affected_village_data:
                # Convert existing population to integer and increment
                try:
                    current_population = int(record.population) if record.population.isdigit() else 0
                    new_population = current_population + population_increment
                    record.population = str(new_population)  
                    record.save()
                    updated_records.append({
                        'pin_code': record.pin_code if record.pin_code else None,
                        'popup': record.popup,
                        'center': record.center,
                        'zoom': record.zoom,
                        'marker': record.marker,
                        'radius': record.radius,
                        'severity': record.severity,
                        'population': record.population,
                        'village_id':record.village.id,
                        'village_name':record.village.display_name

                    })
                except ValueError:
                    return Response(
                        {"error": f"Invalid population value for record with popup: {record.popup}"},
                        status=status.HTTP_400_BAD_REQUEST
                    )

            return Response(
                {
                    "message": f"Successfully updated population for {len(updated_records)} record(s)",
                    "updated_records": updated_records
                },
                status=status.HTTP_200_OK
            )

        except ObjectDoesNotExist:
            return Response(
                {"error": f"No records found for village: {village_id}"},
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            return Response(
                {"error": f"An unexpected error occurred: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class ShowPersonOnUserPage(APIView):
    pagination_class = TempPersonPagination

    def get(self, request):
        search_query = request.query_params.get('search', '')

        queryset = PersonData.objects.all()

        # Filter by search query if provided
        if search_query:
            queryset = queryset.filter(
                models.Q(village__pin_code__icontains=search_query) |
                models.Q(village__display_name__icontains=search_query) |
                models.Q(village__tehsil__name__icontains=search_query) |
                models.Q(village__tehsil__city__name__icontains=search_query) |
                models.Q(village__tehsil__city__state__name__icontains=search_query)
            )

        # Apply pagination
        paginator = self.pagination_class()
        page = paginator.paginate_queryset(queryset, request)

        data = []

        for person_data in page:  # Iterate over paginated queryset
            verification_data = None  

            if person_data.is_single_user:
                try:
                    verified_by = person_data.form.verified_by
                    verification_data = [{
                        'verified_by_data': [{
                            'verification_id': verified_by.id,
                            'verification_image': urljoin(
                                request.build_absolute_uri('/'),
                                verified_by.verification_image.url
                            ) if verified_by.verification_image else None,
                            'surveyor_name': verified_by.surveyor_name,
                            'surveyor_mobile': verified_by.surveyor_mobile,
                            'date': verified_by.date,
                            'is_verified': verified_by.is_verified,
                        }]
                    }]
                except AttributeError:
                    verification_data = None

            data.append({
                'person_id': person_data.id,
                'is_active': person_data.is_active,
                'is_single_user': person_data.is_single_user,
                "village_data": [{
                    'village_id': person_data.village.id,
                    'display_name': person_data.village.display_name,
                    'pin_code': person_data.village.pin_code,
                    'longitude': person_data.village.longitude,
                    'latitude': person_data.village.latitude,
                    "tehsil_data": [{
                        'tehsil_id': person_data.village.tehsil.id,
                        'name': person_data.village.tehsil.name,
                        'city_data': [{
                            'city_id': person_data.village.tehsil.city.id,
                            'name': person_data.village.tehsil.city.name,
                            'state_data': [{
                                'state_id': person_data.village.tehsil.city.state.id,
                                'name': person_data.village.tehsil.city.state.name
                            }]
                        }]
                    }]
                }],
                "form_data": [{
                    "form_id": person_data.form.id,
                    'farmer_name': person_data.form.farmerName,
                    'farmer_image': urljoin(
                        request.build_absolute_uri('/'),
                        person_data.form.farmer_image.url
                    ) if person_data.form.farmer_image else None,
                    'mobileNumber': person_data.form.mobileNumber,
                    'fatherName': person_data.form.fatherName,
                    'email': person_data.form.email,
                    'houseStatus': person_data.form.houseStatus,
                    'totalLandOwned': person_data.form.totalLandOwned,
                    'landAffected': person_data.form.landAffected,
                    'cropsPlanted': person_data.form.cropsPlanted,
                    'cropsLost': person_data.form.cropsLost,
                    'estimatedCropLoss': person_data.form.estimatedCropLoss,
                    'tractorLeveling': person_data.form.tractorLeveling,
                    'manureFertilizer': person_data.form.manureFertilizer,
                    'seedsRequired': person_data.form.seedsRequired,
                    'fertilizersPesticides': person_data.form.fertilizersPesticides,
                    'laborRequirement': person_data.form.laborRequirement,
                    'irrigationRepair': person_data.form.irrigationRepair,
                    'livestockDamage': person_data.form.livestockDamage,
                    'householdNeeds': person_data.form.householdNeeds,
                    'housingRepair': person_data.form.housingRepair,
                    'otherSupport': person_data.form.otherSupport,
                    'additionalNotes': person_data.form.additionalNotes,
                    'amount_needed':person_data.form.amount_needed if person_data.form.amount_needed else None,
                    'amount_received':person_data.form.amount_received if person_data.form.amount_received else None
                }],
                'verification_data': verification_data
            })

        return paginator.get_paginated_response(data)




# class ShowPersonOnUserPageWithFilters(APIView):
#     pagination_class = TempPersonPagination

#     def get(self, request):
#         # Get query parameters
#         search_query = request.query_params.get('search', '')
#         farmer_name = request.query_params.get('farmer_name', '')
#         house_status = request.query_params.get('house_status', '')
#         estimated_crop_loss = request.query_params.get('estimatedCropLoss', '')

#         queryset = PersonData.objects.all().order_by('-created_at')

#         # Search filter for farmer name, father name, mobile number, email, and location fields
#         if search_query:
#             queryset = queryset.filter(
#                 models.Q(form__farmerName__icontains=search_query) |
#                 models.Q(form__fatherName__icontains=search_query) |
#                 models.Q(form__mobileNumber__icontains=search_query) |
#                 models.Q(form__email__icontains=search_query) |
#                 models.Q(village__pin_code__icontains=search_query) |
#                 models.Q(village__display_name__icontains=search_query) |
#                 models.Q(village__tehsil__name__icontains=search_query) |
#                 models.Q(village__tehsil__city__name__icontains=search_query) |
#                 models.Q(village__tehsil__city__state__name__icontains=search_query)
#             )

#         # Search filter for farmer name (supports multiple comma-separated values)
#         if farmer_name:
#             farmer_name_list = [name.strip() for name in farmer_name.split(',')]
#             queryset = queryset.filter(
#                 models.Q(form__farmerName__in=farmer_name_list)
#             )

#         # Filter by house_status (supports multiple comma-separated values)
#         if house_status:
#             house_status_list = [status.strip() for status in house_status.split(',')]
#             queryset = queryset.filter(form__houseStatus__in=house_status_list)

#         # Filter by estimatedCropLoss range (e.g., estimatedCropLoss=1000-20000)
#         if estimated_crop_loss:
#             try:
#                 min_loss, max_loss = map(float, estimated_crop_loss.split('-'))
#                 queryset = queryset.filter(
#                     form__estimatedCropLoss__gte=min_loss,
#                     form__estimatedCropLoss__lte=max_loss
#                 )
#             except ValueError:
#                 # Handle invalid range format gracefully
#                 pass

#         # Apply pagination
#         paginator = self.pagination_class()
#         page = paginator.paginate_queryset(queryset, request)
#         data = []

#         for person_data in page:
#             verification_data = None
#             if person_data.is_single_user:
#                 try:
#                     verified_by = person_data.form.verified_by
#                     verification_data = [{
#                         'verified_by_data': [{
#                             'verification_id': verified_by.id,
#                             'verification_image': urljoin(
#                                 request.build_absolute_uri('/'),
#                                 verified_by.verification_image.url
#                             ) if verified_by.verification_image else None,
#                             'surveyor_name': verified_by.surveyor_name,
#                             'surveyor_mobile': verified_by.surveyor_mobile,
#                             'date': verified_by.date,
#                             'is_verified': verified_by.is_verified,
#                         }]
#                     }]
#                 except AttributeError:
#                     verification_data = None

#             data.append({
#                 'person_id': person_data.id,
#                 'is_active': person_data.is_active,
#                 'is_single_user': person_data.is_single_user,
#                 "village_data": [{
#                     'village_id': person_data.village.id,
#                     'display_name': person_data.village.display_name,
#                     'pin_code': person_data.village.pin_code,
#                     'longitude': person_data.village.longitude,
#                     'latitude': person_data.village.latitude,
#                     "tehsil_data": [{
#                         'tehsil_id': person_data.village.tehsil.id,
#                         'name': person_data.village.tehsil.name,
#                         'city_data': [{
#                             'city_id': person_data.village.tehsil.city.id,
#                             'name': person_data.village.tehsil.city.name,
#                             'state_data': [{
#                                 'state_id': person_data.village.tehsil.city.state.id,
#                                 'name': person_data.village.tehsil.city.state.name
#                             }]
#                         }]
#                     }]
#                 }],
#                 "form_data": [{
#                     "form_id": person_data.form.id,
#                     'farmer_name': person_data.form.farmerName,
#                     'farmer_image': urljoin(
#                         request.build_absolute_uri('/'),
#                         person_data.form.farmer_image.url
#                     ) if person_data.form.farmer_image else None,
#                     'mobileNumber': person_data.form.mobileNumber,
#                     'fatherName': person_data.form.fatherName,
#                     'email': person_data.form.email,
#                     'houseStatus': person_data.form.houseStatus,
#                     'totalLandOwned': person_data.form.totalLandOwned,
#                     'landAffected': person_data.form.landAffected,
#                     'cropsPlanted': person_data.form.cropsPlanted,
#                     'cropsLost': person_data.form.cropsLost,
#                     'estimatedCropLoss': person_data.form.estimatedCropLoss,
#                     'tractorLeveling': person_data.form.tractorLeveling,
#                     'manureFertilizer': person_data.form.manureFertilizer,
#                     'seedsRequired': person_data.form.seedsRequired,
#                     'fertilizersPesticides': person_data.form.fertilizersPesticides,
#                     'laborRequirement': person_data.form.laborRequirement,
#                     'irrigationRepair': person_data.form.irrigationRepair,
#                     'livestockDamage': person_data.form.livestockDamage,
#                     'householdNeeds': person_data.form.householdNeeds,
#                     'housingRepair': person_data.form.housingRepair,
#                     'otherSupport': person_data.form.otherSupport,
#                     'additionalNotes': person_data.form.additionalNotes,
#                     'amount_needed':person_data.form.amount_needed if person_data.form.amount_needed else None,
#                     'amount_received':person_data.form.amount_received if person_data.form.amount_received else None
#                 }],
#                 'verification_data': verification_data
#             })

#         return paginator.get_paginated_response(data)


class ShowPersonOnUserPageWithFilters(APIView):
    pagination_class = TempPersonPagination

    def get(self, request):
        # Get query parameters
        search_query = request.query_params.get('search', '')
        farmer_name = request.query_params.get('farmer_name', '')
        house_status = request.query_params.get('house_status', '')
        estimated_crop_loss = request.query_params.get('estimatedCropLoss', '')

        queryset = PersonData.objects.all().order_by('-created_at')

        # Search filter for farmer name, father name, mobile number, email, and location fields
        if search_query:
            search_terms = [term.strip() for term in search_query.split(',') if term.strip()]
            if search_terms:
                q = models.Q()
                for term in search_terms:
                    q |= (
                        models.Q(form__farmerName__icontains=term) |
                        models.Q(form__fatherName__icontains=term) |
                        models.Q(form__mobileNumber__icontains=term) |
                        models.Q(form__email__icontains=term) |
                        models.Q(village__pin_code__icontains=term) |
                        models.Q(village__display_name__icontains=term) |
                        models.Q(village__tehsil__name__icontains=term) |
                        models.Q(village__tehsil__city__name__icontains=term) 
                    )
                queryset = queryset.filter(q)
            else:
                # If search_query is empty or only contains invalid terms, return empty queryset
                queryset = queryset.none()

        # Search filter for farmer name (supports multiple comma-separated values with icontains)
        if farmer_name:
            farmer_name_list = [name.strip() for name in farmer_name.split(',') if name.strip()]
            if farmer_name_list:
                q = models.Q()
                for name in farmer_name_list:
                    q |= models.Q(form__farmerName__icontains=name)
                queryset = queryset.filter(q)
            else:
                # If farmer_name is empty or invalid, return empty queryset
                queryset = queryset.none()

        # Filter by house_status (supports multiple comma-separated values with icontains)
        if house_status:
            house_status_list = [status.strip() for status in house_status.split(',') if status.strip()]
            if house_status_list:
                q = models.Q()
                for status in house_status_list:
                    q |= models.Q(form__houseStatus__icontains=status)
                queryset = queryset.filter(q)
            else:
                # If house_status is empty or invalid, return empty queryset
                queryset = queryset.none()

        # Filter by estimatedCropLoss range (e.g., estimatedCropLoss=1000-20000)
        if estimated_crop_loss:
            try:
                min_loss, max_loss = map(float, estimated_crop_loss.split('-'))
                queryset = queryset.filter(
                    form__estimatedCropLoss__gte=min_loss,
                    form__estimatedCropLoss__lte=max_loss
                )
            except ValueError:
                # Handle invalid range format by returning empty queryset
                queryset = queryset.none()

        # Apply pagination
        paginator = self.pagination_class()
        page = paginator.paginate_queryset(queryset, request)
        data = []

        for person_data in page:
            verification_data = None
            if person_data.is_single_user:
                try:
                    verified_by = person_data.form.verified_by
                    verification_data = [{
                        'verified_by_data': [{
                            'verification_id': verified_by.id,
                            'verification_image': urljoin(
                                request.build_absolute_uri('/'),
                                verified_by.verification_image.url
                            ) if verified_by.verification_image else None,
                            'surveyor_name': verified_by.surveyor_name,
                            'surveyor_mobile': verified_by.surveyor_mobile,
                            'date': verified_by.date,
                            'is_verified': verified_by.is_verified,
                        }]
                    }]
                except AttributeError:
                    verification_data = None

            data.append({
                'person_id': person_data.id,
                'is_active': person_data.is_active,
                'is_single_user': person_data.is_single_user,
                'created_at':person_data.created_at,
                "village_data": [{
                    'village_id': person_data.village.id,
                    'display_name': person_data.village.display_name,
                    'pin_code': person_data.village.pin_code,
                    'longitude': person_data.village.longitude,
                    'latitude': person_data.village.latitude,
                    "tehsil_data": [{
                        'tehsil_id': person_data.village.tehsil.id,
                        'name': person_data.village.tehsil.name,
                        'city_data': [{
                            'city_id': person_data.village.tehsil.city.id,
                            'name': person_data.village.tehsil.city.name,
                            'state_data': [{
                                'state_id': person_data.village.tehsil.city.state.id,
                                'name': person_data.village.tehsil.city.state.name
                            }]
                        }]
                    }]
                }],
                "form_data": [{
                    "form_id": person_data.form.id,
                    'farmer_name': person_data.form.farmerName,
                    'farmer_image': urljoin(
                        request.build_absolute_uri('/'),
                        person_data.form.farmer_image.url
                    ) if person_data.form.farmer_image else None,
                    'mobileNumber': person_data.form.mobileNumber,
                    'fatherName': person_data.form.fatherName,
                    'email': person_data.form.email,
                    'houseStatus': person_data.form.houseStatus,
                    'totalLandOwned': person_data.form.totalLandOwned,
                    'landAffected': person_data.form.landAffected,
                    'cropsPlanted': person_data.form.cropsPlanted,
                    'cropsLost': person_data.form.cropsLost,
                    'estimatedCropLoss': person_data.form.estimatedCropLoss,
                    'tractorLeveling': person_data.form.tractorLeveling,
                    'manureFertilizer': person_data.form.manureFertilizer,
                    'seedsRequired': person_data.form.seedsRequired,
                    'fertilizersPesticides': person_data.form.fertilizersPesticides,
                    'laborRequirement': person_data.form.laborRequirement,
                    'irrigationRepair': person_data.form.irrigationRepair,
                    'livestockDamage': person_data.form.livestockDamage,
                    'householdNeeds': person_data.form.householdNeeds,
                    'housingRepair': person_data.form.housingRepair,
                    'otherSupport': person_data.form.otherSupport,
                    'additionalNotes': person_data.form.additionalNotes,
                    'amount_needed': person_data.form.amount_needed if person_data.form.amount_needed else None,
                    'amount_received': person_data.form.amount_received if person_data.form.amount_received else None
                }],
                'verification_data': verification_data
            })

        return paginator.get_paginated_response(data)
    

class ExportToExcelAPIView(APIView):
    def get(self, request):
        try:
            # Define column headers
            headers = [
                'farmerName',
                'fatherName',
                'mobileNumber',
                'email',
                'state_id',
                'city_id',
                'tehsil_id',
                'village_id',
                'houseStatus',
                'totalLandOwned',
                'landAffected',
                'cropsPlanted',
                'cropsLost',
                'estimatedCropLoss',
                'tractorLeveling',
                'manureFertilizer',
                'seedsRequired',
                'fertilizersPesticides',
                'laborRequirement',
                'irrigationRepair',
                'livestockDamage',
                'householdNeeds',
                'housingRepair',
                'otherSupport',
                'surveyorName',
                'AmountNeeded'
            ]
            
            # Create workbook and worksheet
            wb = Workbook()
            ws = wb.active
            ws.title = 'Farmers Data'
            
            # Add headers (row 1) and lock them
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.protection = Protection(locked=True)  # Explicitly lock header cells
            
            # Auto-fit column widths based on header length
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Enable sheet protection
            ws.protection = SheetProtection(sheet=True, objects=False, scenarios=False)
            
            # Unlock data cells (rows 2+ for all columns under headers)
            num_cols = len(headers)
            for r in range(2, 1002):  # Unlock first 1000 rows for potential data entry
                for c in range(1, num_cols + 1):
                    ws.cell(row=r, column=c).protection = Protection(locked=False)  # Unlock data cells
            
            # Create excel_reports directory if it doesn't exist
            excel_dir = os.path.join(settings.MEDIA_ROOT, 'excel_reports')
            os.makedirs(excel_dir, exist_ok=True)
            
            # Generate unique filename with timestamp
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f'farmers_data_{timestamp}.xlsx'
            file_path = os.path.join(excel_dir, filename)
            
            # Save the workbook to the file
            wb.save(file_path)
            
            # Construct absolute URL
            relative_url = f'{settings.MEDIA_URL}excel_reports/{filename}'
            absolute_url = request.build_absolute_uri(relative_url)
            
            # Return JSON response with the absolute URL
            response_data = {
                'message': 'Excel file with headers generated successfully',
                'download_url': absolute_url,
                'filename': filename
            }
            
            return Response(response_data, status=status.HTTP_200_OK)
        
        except Exception as e:
            import traceback
            print("❌ ERROR: Failed to generate Excel file")
            print(traceback.format_exc())
            return Response(
                {'error': f'Failed to generate Excel file: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )




# class BulkUploadExcelAPIViewAdmin(APIView):
#     def post(self, request, *args, **kwargs):
#         print("=" * 60)
#         print("🚀 BULK UPLOAD EXCEL APIVIEW - POST REQUEST STARTED")
#         print("=" * 60)
#         print(f"📋 Request method: {request.method}")
#         print(f"📋 Request content type: {request.content_type}")
#         print(f"📋 Request FILES keys: {list(request.FILES.keys())}")
        
#         try:
#             # Check if Excel file is provided
#             excel_file = request.FILES.get('excel_file')
#             if not excel_file:
#                 print("❌ ERROR: No Excel file provided")
#                 return Response({'error': 'Excel file is required'}, status=status.HTTP_400_BAD_REQUEST)
            
#             print(f"📄 Excel file received: {excel_file.name}")
            
#             # Load the workbook
#             wb = load_workbook(excel_file, data_only=True)
#             ws = wb.active
#             if not ws:
#                 print("❌ ERROR: No active worksheet found")
#                 return Response({'error': 'No active worksheet found in Excel file'}, status=status.HTTP_400_BAD_REQUEST)
            
#             # Expected headers
#             expected_headers = [
#                 'farmerName', 'fatherName', 'mobileNumber', 'email', 'state_id', 'city_id',
#                 'tehsil_id', 'village_id', 'houseStatus', 'totalLandOwned',
#                 'landAffected', 'cropsPlanted', 'cropsLost', 'estimatedCropLoss',
#                 'tractorLeveling', 'manureFertilizer', 'seedsRequired', 'fertilizersPesticides',
#                 'laborRequirement', 'irrigationRepair', 'livestockDamage', 'householdNeeds',
#                 'housingRepair', 'otherSupport', 'surveyorName', 'AmountNeeded'
#             ]
            
#             # Validate headers
#             headers = [cell.value for cell in ws[1] if cell.value]
#             if not headers or headers != expected_headers:
#                 print(f"❌ ERROR: Invalid headers. Expected: {expected_headers}, Got: {headers}")
#                 return Response(
#                     {'error': f'Invalid headers in Excel file. Expected: {expected_headers}'},
#                     status=status.HTTP_400_BAD_REQUEST
#                 )
            
#             print("✅ Headers validated successfully")
            
#             # Start database transaction
#             with transaction.atomic():
#                 print("🔒 Database transaction STARTED")
                
#                 # Pre-fetch location data to validate IDs
#                 state_ids = set(State.objects.values_list('id', flat=True))
#                 city_ids = set(City.objects.values_list('id', flat=True))
#                 tehsil_ids = set(Tehsil.objects.values_list('id', flat=True))
#                 village_ids = set(Village.objects.values_list('id', flat=True))

#                 # Collect data for bulk creation
#                 verified_by_data = []
#                 farmer_form_data = []
#                 user_data = []
#                 person_data = []
#                 created_forms = []

#                 existing_usernames = set(User.objects.values_list('username', flat=True))  # Cache existing usernames

#                 for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
#                     if all(cell is None for cell in row):  # Skip empty rows
#                         print(f"⚠️ Skipping empty row {row_num}")
#                         continue
                    
#                     print(f"\n🔄 Processing row {row_num} (Transaction Active)")
#                     data = dict(zip(expected_headers, row))
#                     print(f"👤 Farmer Name: {data.get('farmerName', 'N/A')}")
#                     print(f"📱 Mobile: {data.get('mobileNumber', 'N/A')}")
                    
#                     # Validate required fields
#                     required_fields = ['farmerName', 'fatherName', 'mobileNumber', 'state_id', 'city_id', 'tehsil_id', 'village_id', 'surveyorName', 'AmountNeeded']
#                     missing_fields = [field for field in required_fields if not data.get(field)]
#                     if missing_fields:
#                         print(f"❌ ERROR: Missing required fields in row {row_num}: {missing_fields}")
#                         raise ValueError(f'Missing required fields in row {row_num}: {missing_fields}')
                    
#                     # Validate IDs
#                     state_id = int(data['state_id'])
#                     city_id = int(data['city_id'])
#                     tehsil_id = int(data['tehsil_id'])
#                     village_id = int(data['village_id'])
                    
#                     if state_id not in state_ids:
#                         raise ValueError(f"Invalid state_id {state_id} in row {row_num}")
#                     if city_id not in city_ids:
#                         raise ValueError(f"Invalid city_id {city_id} in row {row_num}")
#                     if tehsil_id not in tehsil_ids:
#                         raise ValueError(f"Invalid tehsil_id {tehsil_id} in row {row_num}")
#                     if village_id not in village_ids:
#                         raise ValueError(f"Invalid village_id {village_id} in row {row_num}")
                    
#                     # Fetch location objects
#                     state = State.objects.get(id=state_id)
#                     city = City.objects.get(id=city_id, state=state)
#                     tehsil = Tehsil.objects.get(id=tehsil_id, city=city)
#                     village = Village.objects.get(id=village_id, tehsil=tehsil)
                    
#                     print("\n🌐 Location Hierarchy validated:")
#                     print(f"   ✅ State: ID {state.id}, Name {state.name}")
#                     print(f"   ✅ City: ID {city.id}, Name {city.name}")
#                     print(f"   ✅ Tehsil: ID {tehsil.id}, Name {tehsil.name}")
#                     print(f"   ✅ Village: ID {village.id}, Name {village.display_name}")
                    
#                     # Create VerifiedBy data
#                     verified_by = VerifiedBy(
#                         verification_image=None,
#                         surveyor_name=data['surveyorName'],
#                         date=datetime.now().date(),
#                         surveyor_mobile='',
#                         is_verified=True
#                     )
#                     verified_by_data.append(verified_by)
#                     print(f"   📸 VerifiedBy prepared for row {row_num}")
                    
#                     # Create FarmerForm data
#                     farmer_form = FarmerForm(
#                         verified_by=verified_by,  # Will be set after bulk_create
#                         farmer_image=None,
#                         farmerName=data['farmerName'],
#                         fatherName=data['fatherName'],
#                         mobileNumber=data['mobileNumber'],
#                         email=data.get('email', ''),
#                         houseStatus=data.get('houseStatus', ''),
#                         totalLandOwned=data.get('totalLandOwned', ''),
#                         landAffected=data.get('landAffected', ''),
#                         cropsPlanted=data.get('cropsPlanted', ''),
#                         cropsLost=data.get('cropsLost', ''),
#                         estimatedCropLoss=data.get('estimatedCropLoss', ''),
#                         tractorLeveling=data.get('tractorLeveling', ''),
#                         manureFertilizer=data.get('manureFertilizer', ''),
#                         seedsRequired=data.get('seedsRequired', ''),
#                         fertilizersPesticides=data.get('fertilizersPesticides', ''),
#                         laborRequirement=data.get('laborRequirement', ''),
#                         irrigationRepair=data.get('irrigationRepair', ''),
#                         livestockDamage=data.get('livestockDamage', ''),
#                         householdNeeds=data.get('householdNeeds', ''),
#                         housingRepair=data.get('housingRepair', ''),
#                         otherSupport=data.get('otherSupport', ''),
#                         additionalNotes=data.get('additionalNotes', ''),
#                         is_active=True,
#                         amount_needed=data.get('AmountNeeded', 0)
#                     )
#                     farmer_form_data.append(farmer_form)
#                     print(f"   👨‍🌾 FarmerForm prepared for row {row_num}")
                    
#                     # Create User data with unique username
#                     email = data.get('email', '')
#                     username_base = email if email else f"{data['farmerName'].replace(' ', '_').lower()}"
#                     username = username_base
#                     counter = 1
#                     while username in existing_usernames or any(u.username == username for u in user_data):
#                         username = f"{username_base}_{counter}"
#                         counter += 1
                    
#                     existing_usernames.add(username)  # Add to set to track for this batch
                    
#                     user = User(
#                         username=username,
#                         email=email if email else '',
#                         password='abc123',  # Temporary password
#                         first_name=data['farmerName'].split()[0] if data['farmerName'] else '',
#                         last_name=' '.join(data['farmerName'].split()[1:]) if data['farmerName'] and len(data['farmerName'].split()) > 1 else ''
#                     )
#                     user_data.append(user)
#                     print(f"   👤 User prepared for row {row_num}: {username}")
                    
#                     # Create PersonData data
#                     person = PersonData(
#                         village=village,
#                         form=farmer_form,  # Will be set after bulk_create
#                         user=user,
#                         is_active=True,
#                         is_single_user=False
#                     )
#                     person_data.append(person)
#                     print(f"   📝 PersonData prepared for row {row_num}")
                    
#                     # Prepare response data
#                     created_forms.append({
#                         'id': farmer_form.id,  # Will be set after bulk_create
#                         'farmerName': data['farmerName'],
#                         'mobileNumber': data['mobileNumber'],
#                         'village': village.display_name,
#                         'user': {
#                             'id': user.id,  # Will be set after bulk_create
#                             'username': user.username,
#                             'email': user.email if user.email else ''
#                         },
#                         'person_data_id': person.id,  # Will be set after bulk_create
#                         'created_at': datetime.now().isoformat()  # Approximate, updated later
#                     })
#                     print(f"✅ Row {row_num} prepared successfully!")
#                     print("-" * 40)
                
#                 # Bulk create all objects
#                 print("🔄 Performing bulk create operations...")
#                 verified_by_objects = VerifiedBy.objects.bulk_create(verified_by_data)
#                 farmer_form_objects = FarmerForm.objects.bulk_create(farmer_form_data)
#                 user_objects = User.objects.bulk_create(user_data)
#                 person_data_objects = PersonData.objects.bulk_create(person_data)

#                 # Update foreign key references and IDs
#                 for i, (verified_by, farmer_form, user, person) in enumerate(zip(
#                     verified_by_objects, farmer_form_objects, user_objects, person_data_objects
#                 )):
#                     farmer_form.verified_by = verified_by
#                     farmer_form.save(update_fields=['verified_by'])
#                     person.form = farmer_form
#                     person.user = user
#                     person.save(update_fields=['form', 'user'])
#                     created_forms[i]['id'] = farmer_form.id
#                     created_forms[i]['user']['id'] = user.id
#                     created_forms[i]['person_data_id'] = person.id
#                     created_forms[i]['created_at'] = farmer_form.created_at.isoformat()

#                 print(f"\n🎉 All {len(created_forms)} rows processed successfully!")
#                 print("🔓 Database transaction COMMITTED")
                
#                 response_data = {
#                     'message': 'Excel data processed successfully',
#                     'created_forms': created_forms,
#                     'total_processed': len(created_forms),
#                     'transaction_status': 'committed',
#                     'users_created': len(created_forms),
#                     'person_data_created': len(created_forms)
#                 }
                
#                 print("📤 Sending success response...")
#                 return Response(response_data, status=status.HTTP_201_CREATED)
            
#         except ValueError as e:
#             print(f"❌ VALUE ERROR: {str(e)}")
#             print("🔓 Database transaction ROLLEDBACK due to validation error")
#             return Response({'error': f'Validation error: {str(e)}'}, status=status.HTTP_400_BAD_REQUEST)
        
#         except Exception as e:
#             print(f"❌ GENERAL EXCEPTION: {str(e)}")
#             print("🔓 Database transaction ROLLEDBACK due to exception")
#             traceback.print_exc()
#             return Response({'error': f'Server error: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
#         finally:
#             print("=" * 60)
#             print("🏁 BULK UPLOAD EXCEL APIVIEW - POST REQUEST COMPLETED")
#             print("=" * 60)



class BulkUploadExcelAPIViewAdmin(APIView):
    def post(self, request, *args, **kwargs):
        print("=" * 60)
        print("🚀 BULK UPLOAD EXCEL APIVIEW - POST REQUEST STARTED")
        print("=" * 60)
        print(f"📋 Request method: {request.method}")
        print(f"📋 Request content type: {request.content_type}")
        print(f"📋 Request FILES keys: {list(request.FILES.keys())}")
        
        try:
            # Check if Excel file is provided
            excel_file = request.FILES.get('excel_file')
            if not excel_file:
                print("❌ ERROR: No Excel file provided")
                return Response({'error': 'Excel file is required'}, status=status.HTTP_400_BAD_REQUEST)
            
            print(f"📄 Excel file received: {excel_file.name}")
            
            # Load the workbook
            wb = load_workbook(excel_file, data_only=True)
            ws = wb.active
            if not ws:
                print("❌ ERROR: No active worksheet found")
                return Response({'error': 'No active worksheet found in Excel file'}, status=status.HTTP_400_BAD_REQUEST)
            
            # Expected headers
            expected_headers = [
                'farmerName', 'fatherName', 'mobileNumber', 'email', 'state_id', 'city_id',
                'tehsil_id', 'village_id', 'houseStatus', 'totalLandOwned',
                'landAffected', 'cropsPlanted', 'cropsLost', 'estimatedCropLoss',
                'tractorLeveling', 'manureFertilizer', 'seedsRequired', 'fertilizersPesticides',
                'laborRequirement', 'irrigationRepair', 'livestockDamage', 'householdNeeds',
                'housingRepair', 'otherSupport', 'surveyorName', 'AmountNeeded'
            ]
            
            # Validate headers
            headers = [cell.value for cell in ws[1] if cell.value]
            if not headers or headers != expected_headers:
                print(f"❌ ERROR: Invalid headers. Expected: {expected_headers}, Got: {headers}")
                return Response(
                    {'error': f'Invalid headers in Excel file. Expected: {expected_headers}'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            print("✅ Headers validated successfully")
            
            # Pre-fetch location data to validate IDs
            state_ids = set(State.objects.values_list('id', flat=True))
            city_ids = set(City.objects.values_list('id', flat=True))
            tehsil_ids = set(Tehsil.objects.values_list('id', flat=True))
            village_ids = set(Village.objects.values_list('id', flat=True))

            # Collect data for bulk creation
            verified_by_data = []
            farmer_form_data = []
            user_data = []
            person_data = []
            created_forms = []
            errors = []

            existing_usernames = set(User.objects.values_list('username', flat=True))  # Cache existing usernames

            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
                if all(cell is None for cell in row):  # Skip empty rows
                    print(f"⚠️ Skipping empty row {row_num}")
                    continue
                
                print(f"\n🔄 Processing row {row_num}")
                try:
                    data = dict(zip(expected_headers, row))
                    print(f"👤 Farmer Name: {data.get('farmerName', 'N/A')}")
                    print(f"📱 Mobile: {data.get('mobileNumber', 'N/A')}")
                    
                    # Validate required fields
                    required_fields = ['farmerName', 'fatherName', 'mobileNumber', 'state_id', 'city_id', 'tehsil_id', 'village_id', 'surveyorName', 'AmountNeeded']
                    missing_fields = [field for field in required_fields if not data.get(field)]
                    if missing_fields:
                        raise ValueError(f'Missing required fields: {missing_fields}')
                    
                    # Validate IDs
                    state_id = int(data['state_id'])
                    city_id = int(data['city_id'])
                    tehsil_id = int(data['tehsil_id'])
                    village_id = int(data['village_id'])
                    
                    if state_id not in state_ids:
                        raise ValueError(f"Invalid state_id {state_id}")
                    if city_id not in city_ids:
                        raise ValueError(f"Invalid city_id {city_id}")
                    if tehsil_id not in tehsil_ids:
                        raise ValueError(f"Invalid tehsil_id {tehsil_id}")
                    if village_id not in village_ids:
                        raise ValueError(f"Invalid village_id {village_id}")
                    
                    # Fetch location objects
                    state = State.objects.get(id=state_id)
                    city = City.objects.get(id=city_id, state=state)
                    tehsil = Tehsil.objects.get(id=tehsil_id, city=city)
                    village = Village.objects.get(id=village_id, tehsil=tehsil)
                    
                    print("\n🌐 Location Hierarchy validated:")
                    print(f"   ✅ State: ID {state.id}, Name {state.name}")
                    print(f"   ✅ City: ID {city.id}, Name {city.name}")
                    print(f"   ✅ Tehsil: ID {tehsil.id}, Name {tehsil.name}")
                    print(f"   ✅ Village: ID {village.id}, Name {village.display_name}")
                    
                    # Create VerifiedBy data
                    verified_by = VerifiedBy(
                        verification_image=None,
                        surveyor_name=data['surveyorName'],
                        date=datetime.now().date(),
                        surveyor_mobile='',
                        is_verified=True
                    )
                    verified_by_data.append(verified_by)
                    print(f"   📸 VerifiedBy prepared for row {row_num}")
                    
                    # Create FarmerForm data
                    farmer_form = FarmerForm(
                        verified_by=verified_by,  # Will be set after bulk_create
                        farmer_image=None,
                        farmerName=data['farmerName'],
                        fatherName=data['fatherName'],
                        mobileNumber=data['mobileNumber'],
                        email=data.get('email', ''),
                        houseStatus=data.get('houseStatus', ''),
                        totalLandOwned=data.get('totalLandOwned', ''),
                        landAffected=data.get('landAffected', ''),
                        cropsPlanted=data.get('cropsPlanted', ''),
                        cropsLost=data.get('cropsLost', ''),
                        estimatedCropLoss=data.get('estimatedCropLoss', ''),
                        tractorLeveling=data.get('tractorLeveling', ''),
                        manureFertilizer=data.get('manureFertilizer', ''),
                        seedsRequired=data.get('seedsRequired', ''),
                        fertilizersPesticides=data.get('fertilizersPesticides', ''),
                        laborRequirement=data.get('laborRequirement', ''),
                        irrigationRepair=data.get('irrigationRepair', ''),
                        livestockDamage=data.get('livestockDamage', ''),
                        householdNeeds=data.get('householdNeeds', ''),
                        housingRepair=data.get('housingRepair', ''),
                        otherSupport=data.get('otherSupport', ''),
                        additionalNotes=data.get('additionalNotes', ''),
                        is_active=True,
                        amount_needed=data.get('AmountNeeded', 0)
                    )
                    farmer_form_data.append(farmer_form)
                    print(f"   👨‍🌾 FarmerForm prepared for row {row_num}")
                    
                    # Create User data with unique username
                    email = data.get('email', '')
                    username_base = email if email else f"{data['farmerName'].replace(' ', '_').lower()}"
                    username = username_base
                    counter = 1
                    while username in existing_usernames or any(u.username == username for u in user_data):
                        username = f"{username_base}_{counter}"
                        counter += 1
                    
                    user = User(
                        username=username,
                        email=email if email else '',
                        password='abc123',  # Temporary password
                        first_name=data['farmerName'].split()[0] if data['farmerName'] else '',
                        last_name=' '.join(data['farmerName'].split()[1:]) if data['farmerName'] and len(data['farmerName'].split()) > 1 else ''
                    )
                    user_data.append(user)
                    print(f"   👤 User prepared for row {row_num}: {username}")
                    
                    # Create PersonData data
                    person = PersonData(
                        village=village,
                        form=farmer_form,  # Will be set after bulk_create
                        user=user,
                        is_active=True,
                        is_single_user=False
                    )
                    person_data.append(person)
                    print(f"   📝 PersonData prepared for row {row_num}")
                    
                    # Prepare response data
                    created_forms.append({
                        'id': farmer_form.id,  # Will be set after bulk_create
                        'farmerName': data['farmerName'],
                        'mobileNumber': data['mobileNumber'],
                        'village': village.display_name,
                        'user': {
                            'id': user.id,  # Will be set after bulk_create
                            'username': user.username,
                            'email': user.email if user.email else ''
                        },
                        'person_data_id': person.id,  # Will be set after bulk_create
                        'created_at': datetime.now().isoformat()  # Approximate, updated later
                    })
                    
                    existing_usernames.add(username)  # Add only if successful
                    
                    print(f"✅ Row {row_num} prepared successfully!")
                    print("-" * 40)
                
                except Exception as e:
                    print(f"❌ ERROR in row {row_num}: {str(e)}")
                    errors.append({
                        'row_num': row_num,
                        'data': data if 'data' in locals() else {},
                        'error': str(e)
                    })
                    continue
            
            # Bulk create all valid objects if any
            if farmer_form_data:
                with transaction.atomic():
                    print("🔒 Database transaction STARTED")
                    print("🔄 Performing bulk create operations...")
                    verified_by_objects = VerifiedBy.objects.bulk_create(verified_by_data)
                    farmer_form_objects = FarmerForm.objects.bulk_create(farmer_form_data)
                    user_objects = User.objects.bulk_create(user_data)
                    person_data_objects = PersonData.objects.bulk_create(person_data)

                    # Update foreign key references and IDs
                    for i, (verified_by, farmer_form, user, person) in enumerate(zip(
                        verified_by_objects, farmer_form_objects, user_objects, person_data_objects
                    )):
                        farmer_form.verified_by = verified_by
                        farmer_form.save(update_fields=['verified_by'])
                        person.form = farmer_form
                        person.user = user
                        person.save(update_fields=['form', 'user'])
                        created_forms[i]['id'] = farmer_form.id
                        created_forms[i]['user']['id'] = user.id
                        created_forms[i]['person_data_id'] = person.id
                        created_forms[i]['created_at'] = farmer_form.created_at.isoformat()

                    print("🔓 Database transaction COMMITTED")
            
            print(f"\n🎉 {len(created_forms)} rows processed successfully!")
            
            # Generate error Excel if there are errors
            error_file = None
            if errors:
                from openpyxl import Workbook
                error_wb = Workbook()
                error_ws = error_wb.active
                error_ws.title = "Error Records"
                
                error_headers = expected_headers + ['row_number', 'error_message']
                error_ws.append(error_headers)
                
                for err in errors:
                    row_data = [err['data'].get(header, '') for header in expected_headers]
                    error_ws.append(row_data + [err['row_num'], err['error']])
                
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                error_filename = f"error_farmers_data_{timestamp}.xlsx"
                excel_dir = os.path.join(settings.MEDIA_ROOT, 'excel_reports')
                os.makedirs(excel_dir, exist_ok=True)
                file_path = os.path.join(excel_dir, error_filename)
                
                # Save the workbook to the file
                error_wb.save(file_path)
                
                # Construct absolute URL
                relative_url = f'{settings.MEDIA_URL}excel_reports/{error_filename}'
                error_file_url = request.build_absolute_uri(relative_url)
                
                print(f"📄 Generated error Excel: {error_filename} at {error_file_url}")
            
            response_data = {
                'message': 'Excel data processed successfully' if not errors else 'Excel data processed with some errors',
                'created_forms': created_forms,
                'total_processed': len(created_forms),
                'transaction_status': 'committed' if farmer_form_data else 'no valid data',
                'users_created': len(created_forms),
                'person_data_created': len(created_forms),
                'errors_count': len(errors),
                'error_file_url': error_file_url
            }
            
            print("📤 Sending response...")
            return Response(response_data, status=status.HTTP_201_CREATED)
        
        except Exception as e:
            print(f"❌ GENERAL EXCEPTION: {str(e)}")
            traceback.print_exc()
            return Response({'error': f'Server error: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
        finally:
            print("=" * 60)
            print("🏁 BULK UPLOAD EXCEL APIVIEW - POST REQUEST COMPLETED")
            print("=" * 60)

class ShowPersonStats(APIView):
    def get(self, request):
        # Get search query parameter
        search_query = request.query_params.get('search', '')

        # Base queryset for PersonData
        queryset = PersonData.objects.all()

        # Apply location-based search filters
        if search_query:
            queryset = queryset.filter(
                models.Q(village__pin_code__icontains=search_query) |
                models.Q(village__display_name__icontains=search_query) |
                models.Q(village__tehsil__name__icontains=search_query) |
                models.Q(village__tehsil__city__name__icontains=search_query) |
                models.Q(village__tehsil__city__state__name__icontains=search_query)
            )

        # Get related FarmerForm data
        person_data = queryset.select_related('form')

        # Check if any data exists
        if not person_data.exists():
            return Response([])  # Return empty array if no data

        # 1. Affected Person Count
        affected_person_count = person_data.count()

        # 2. Estimated Crop Loss (aggregate as range text)
        crop_loss_ranges = person_data.values_list('form__estimatedCropLoss', flat=True).distinct()
        crop_loss_summary = ", ".join(sorted({x for x in crop_loss_ranges if x}))

        # 3. Count of House Status (Partially Damaged, Fully Damaged)
        house_status_counts = person_data.values('form__houseStatus').annotate(count=models.Count('id'))
        partially_damaged_count = sum(1 for item in house_status_counts if item['form__houseStatus'] == 'Partially Damaged')
        fully_damaged_count = sum(1 for item in house_status_counts if item['form__houseStatus'] == 'Fully Damaged')

        # 4. Sum of Total Land Owned and Land Affected
        total_land_owned_sum = person_data.aggregate(total=models.Sum('form__totalLandOwned'))['total'] or 0
        land_affected_sum = person_data.aggregate(total=models.Sum('form__landAffected'))['total'] or 0

        # 5. Sum of Crops Lost
        crops_lost_sum = person_data.aggregate(total=models.Sum('form__cropsLost'))['total'] or 0

        # Additional Stats
        # 6. Average Land Affected
        avg_land_affected = land_affected_sum / affected_person_count if affected_person_count > 0 else 0

        # 7. Percentage of Affected Land
        percentage_affected_land = (land_affected_sum / total_land_owned_sum * 100) if total_land_owned_sum > 0 else 0

        # Prepare response data
        stats = [{
            'affected_person_count': affected_person_count,
            'estimated_crop_loss_range': crop_loss_summary,
            'partially_damaged_count': partially_damaged_count,
            'fully_damaged_count': fully_damaged_count,
            'total_land_owned_sum': float(total_land_owned_sum) if total_land_owned_sum else 0,
            'land_affected_sum': float(land_affected_sum) if land_affected_sum else 0,
            'crops_lost_sum': float(crops_lost_sum) if crops_lost_sum else 0,
            'avg_land_affected': round(avg_land_affected, 2),
            'percentage_affected_land': round(percentage_affected_land, 2),
        }]

        return Response(stats)



class ProcessedUnprocessedCount(APIView):
    def get(self, request):

        village_id = request.query_params.get('village_id')
        try:
            # Fetch AffectedVillageMapData for the given pin_code
            affected_village_data = AffectedVillageMapData.objects.filter(village__id=village_id)

            # Calculate total population (sum of population field)
            processed_count = sum(int(item.population) for item in affected_village_data if item.population.isdigit())

            affected_village_response = [
                {
                    'pin_code': item.pin_code,
                    'popup': item.popup,
                    'center': item.center,
                    'zoom': item.zoom,
                    'marker': item.marker,
                    'radius': item.radius,
                    'severity': item.severity,
                    'population': item.population,
                    'village_id':item.village.id
                } for item in affected_village_data
            ]

            # Fetch PersonData for Villages with the given pin_code
            person_data = PersonData.objects.filter(village__id=village_id)
            total_person_count = person_data.count()
            person_data_response = [
                {
                    'id': item.id,
                    'farmer_name': item.form.farmerName,
                    'farmer_form_id': item.form.id,
                    'village_display_name': item.village.display_name,
                    'village_id':item.village.id,
                    'village_pin_code': item.village.pin_code,
                    'is_active': item.is_active,
                    'is_single_user': item.is_single_user
                } for item in person_data
            ]

            # Calculate unprocessed count
            unprocessed_count = total_person_count - processed_count

            # Construct the response
            response_data = {
                'affected_village_map_data': affected_village_response,
                'person_data': person_data_response,
                'person_data_count': total_person_count,
                'processed_count': processed_count,
                'unprocessed_count': unprocessed_count
            }

            return Response(response_data, status=status.HTTP_200_OK)

        except (ObjectDoesNotExist, ValueError, ValidationError) as e:
            return Response({'error': f'Invalid data or resource not found: {str(e)}'}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({'error': f'Server error: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class ShowFarmerAmount(APIView):
    def get(self, request):
        # Get person_id from query_params
        person_id = request.query_params.get('person_id')

        # Validate input
        if not person_id:
            return Response(
                {"error": "person_id is required"},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            # Get PersonData object
            person = PersonData.objects.get(id=person_id)

            # Get amount_received and amount_needed, default to '0' if None
            amount_received = person.form.amount_received or '0'
            amount_needed = person.form.amount_needed or '0'

            # Convert to float for calculation
            try:
                amount_received_float = float(amount_received)
                amount_needed_float = float(amount_needed)
                remaining_amount = amount_needed_float - amount_received_float
            except ValueError:
                return Response(
                    {"error": "Invalid amount format: amounts must be numeric"},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Prepare response data
            data = {
                'id': person.id,
                'farmer_id': person.form.id,
                'farmer_name': person.form.farmerName,
                'amount_received': amount_received,
                'amount_needed': amount_needed,
                'remaining_amount': str(remaining_amount)  # Convert back to string for consistency
            }

            return Response(data, status=status.HTTP_200_OK)

        except ObjectDoesNotExist:
            return Response(
                {"error": "PersonData with given ID not found"},
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            return Response(
                {"error": f"An unexpected error occurred: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        
class AddFarmerAmount(APIView):
    def post(self, request):
        # Get parameters from query_params
        person_id = request.query_params.get('person_id')
        amount_received = request.query_params.get('amount_received')

        # Validate input parameters
        if not person_id or not amount_received:
            return Response(
                {"error": "person_id and amount_received are required"},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            # Convert amount_received to Decimal for precise calculation
            try:
                amount_received = Decimal(amount_received)
                if amount_received <= 0:
                    return Response(
                        {"error": "amount_received must be a positive number"},
                        status=status.HTTP_400_BAD_REQUEST
                    )
            except (ValueError, TypeError):
                return Response(
                    {"error": "amount_received must be a valid number"},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Get PersonData object
            person = PersonData.objects.get(id=person_id)
            
            # Get associated FarmerForm
            farmer_form = person.form

            # Convert amount_needed and current amount_received to Decimal
            try:
                amount_needed = Decimal(farmer_form.amount_needed or '0')
                current_amount_received = Decimal(farmer_form.amount_received or '0')
            except (ValueError, TypeError):
                return Response(
                    {"error": "Invalid amount format in FarmerForm"},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Calculate new amount_received
            new_amount_received = current_amount_received + amount_received

            # Validate if new_amount_received doesn't exceed amount_needed
            if new_amount_received > amount_needed:
                return Response(
                    {"error": "Total amount received cannot exceed amount needed"},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Update FarmerForm with new amount_received
            with transaction.atomic():
                farmer_form.amount_received = str(new_amount_received)
                farmer_form.save()

            return Response(
                {
                    "message": "Amount updated successfully",
                    "person_id": person_id,
                    "new_amount_received": str(new_amount_received),
                    "amount_needed": str(amount_needed)
                },
                status=status.HTTP_200_OK
            )

        except ObjectDoesNotExist:
            return Response(
                {"error": "PersonData with given ID not found"},
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            return Response(
                {"error": f"An unexpected error occurred: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )





class GeneratePersonDataPDF(APIView):
    def get(self, request):
        # Get query parameters
        search_query = request.query_params.get('search', '')
        farmer_name = request.query_params.get('farmer_name', '')
        house_status = request.query_params.get('house_status', '')
        estimated_crop_loss = request.query_params.get('estimatedCropLoss', '')

        village_id = request.query_params.get('village_id', '')
        tehsil_id = request.query_params.get('tehsil_id', '')
        city_id = request.query_params.get('district_id', '')

        queryset = PersonData.objects.all()

        if search_query:
            queryset = queryset.filter(
                models.Q(form__farmerName__icontains=search_query) |
                models.Q(form__fatherName__icontains=search_query) |
                models.Q(form__mobileNumber__icontains=search_query) |
                models.Q(form__email__icontains=search_query) |
                models.Q(village__pin_code__icontains=search_query) |
                models.Q(village__display_name__icontains=search_query) |
                models.Q(village__tehsil__name__icontains=search_query) |
                models.Q(village__tehsil__city__name__icontains=search_query) |
                models.Q(village__tehsil__city__state__name__icontains=search_query)
            )

        # Search filter for farmer name (supports multiple comma-separated values)
        if farmer_name:
            farmer_name_list = [name.strip() for name in farmer_name.split(',')]
            queryset = queryset.filter(
                models.Q(form__farmerName__in=farmer_name_list)
            )

        # Filter by house_status (supports multiple comma-separated values)
        if house_status:
            house_status_list = [status.strip() for status in house_status.split(',')]
            queryset = queryset.filter(form__houseStatus__in=house_status_list)

        # Filter by estimatedCropLoss range (e.g., estimatedCropLoss=1000-20000)
        if estimated_crop_loss:
            try:
                min_loss, max_loss = map(float, estimated_crop_loss.split('-'))
                queryset = queryset.filter(
                    form__estimatedCropLoss__gte=min_loss,
                    form__estimatedCropLoss__lte=max_loss
                )
            except ValueError:
                # Handle invalid range format gracefully
                pass

        

        # Filter by village_id (supports multiple comma-separated IDs)
        if village_id:
            village_id_list = [vid.strip() for vid in village_id.split(',')]
            try:
                village_id_list = [int(vid) for vid in village_id_list]  # Convert to integers
                queryset = queryset.filter(village__id__in=village_id_list)
            except ValueError:
                # Handle invalid village_id format gracefully
                pass

        # Filter by tehsil_id (supports multiple comma-separated IDs)
        if tehsil_id:
            tehsil_id_list = [tid.strip() for tid in tehsil_id.split(',')]
            try:
                tehsil_id_list = [int(tid) for tid in tehsil_id_list]  # Convert to integers
                queryset = queryset.filter(village__tehsil__id__in=tehsil_id_list)
            except ValueError:
                # Handle invalid tehsil_id format gracefully
                pass

        # Filter by city_id (supports multiple comma-separated IDs)
        if city_id:
            city_id_list = [cid.strip() for cid in city_id.split(',')]
            try:
                city_id_list = [int(cid) for cid in city_id_list]  # Convert to integers
                queryset = queryset.filter(village__tehsil__city__id__in=city_id_list)
            except ValueError:
                # Handle invalid city_id format gracefully
                pass


        # Prepare data for PDF with SrNo
        data = []
        for index, person_data in enumerate(queryset, 1):
            data.append({
                'SrNo': index,
                'farmerName': person_data.form.farmerName,
                'fatherName': person_data.form.fatherName,
                'mobile_number':person_data.form.mobileNumber,
                'city_name': person_data.village.tehsil.city.name,
                'village_name': person_data.village.display_name,
                'state_name': person_data.village.tehsil.city.state.name,
                'pincode': person_data.village.pin_code,
                'houseStatus': person_data.form.houseStatus,
                'amount_needed': person_data.form.amount_needed or 'N/A',
                'amount_received': person_data.form.amount_received or 'N/A',
                'totalLandOwned': person_data.form.totalLandOwned,
                'landAffected': person_data.form.landAffected,
                'cropsPlanted': person_data.form.cropsPlanted,
                'cropsLost': person_data.form.cropsLost,
                'estimatedCropLoss': person_data.form.estimatedCropLoss,
            })

        # Render the template with data
        rendered_html = get_pdf_template(data, len(data))

        # Generate PDF
        pdf_file_name = f"person_data_report_{uuid.uuid4()}.pdf"
        pdf_file_path = os.path.join(settings.MEDIA_ROOT, 'reports', pdf_file_name)
        os.makedirs(os.path.dirname(pdf_file_path), exist_ok=True)
        HTML(string=rendered_html).write_pdf(pdf_file_path)

        # Generate absolute URL for the PDF
        pdf_url = urljoin(request.build_absolute_uri('/'), os.path.join(settings.MEDIA_URL, 'reports', pdf_file_name))

        return HttpResponse(pdf_url, content_type='text/plain')




class DistrictList(APIView):
    def get(self, request):
        state_id = request.query_params.get('state_id')
        
        if state_id:
            districts = City.objects.filter(state__id=state_id).order_by('created_at')
        else:
            districts = City.objects.all().order_by('created_at')
        
        seen = OrderedDict()
        for district in districts:
            lower_name = district.name.upper()
            if lower_name not in seen:
                seen[lower_name] = {
                    'id': district.id,  
                    'name': lower_name
                }
        
        data = list(seen.values())
        
        return Response(data, status=status.HTTP_200_OK)

class TehsilList(APIView):
    def get(self, request):
        district_name = request.query_params.get('district_name')  
        
        if district_name:
            tehsils = Tehsil.objects.filter(city__name__iexact=district_name).order_by('created_at')
        else:
            tehsils = Tehsil.objects.all().order_by('created_at')
        
        seen = OrderedDict()
        for tehsil in tehsils:
            lower_name = tehsil.name.upper()
            if lower_name not in seen:
                seen[lower_name] = {
                    'id': tehsil.id,  
                    'name': lower_name
                }
        
        data = list(seen.values())
        
        return Response(data, status=status.HTTP_200_OK)

class VillageList(APIView):
    def get(self, request):
        tehsil_name = request.query_params.get('tehsil_name')  
        
        if tehsil_name:
            villages = Village.objects.filter(tehsil__name__iexact=tehsil_name).order_by('created_at')
        else:
            villages = Village.objects.all().order_by('created_at')
        
        seen = OrderedDict()
        for village in villages:
            lower_display = village.display_name.upper()
            if lower_display not in seen:
                seen[lower_display] = {
                    'id': village.id,  
                    'name': lower_display,
                    'longitude':village.longitude,
                    'latitude':village.latitude
                }
        
        data = list(seen.values())
        
        return Response(data, status=status.HTTP_200_OK)



class StateList(APIView):
    def get(self,request):
        state = State.objects.all()

        data = []

        for states in state:
            data.append({
                'state_id':states.id,
                'name':states.name
            })

        
        return Response(data,status=status.HTTP_200_OK)




class VillageListForAddData(APIView):
    def get(self, request):
        # Get unique villages with id and name
        villages = Village.objects.filter(person_village__isnull=False).distinct().values('id', 'display_name','longitude','latitude')
        
        # Convert queryset to list for JSON response
        village_list = [{'id': village['id'], 'name': village['display_name'],'longitude':village['longitude'], 'latitude':village['latitude']} for village in villages]
        
        return Response(village_list, status=status.HTTP_200_OK)






class ShowPersonOnUserPageWithFiltersUsingVillage(APIView):
    pagination_class = TempPersonPagination

    def get(self, request):
        # Get query parameters
        search_query = request.query_params.get('search', '')
        farmer_name = request.query_params.get('farmer_name', '')
        house_status = request.query_params.get('house_status', '')
        estimated_crop_loss = request.query_params.get('estimatedCropLoss', '')
        village_id = request.query_params.get('village_id', '')

        tehsil_id = request.query_params.get('tehsil_id', '')
        city_id = request.query_params.get('city_id', '')

        queryset = PersonData.objects.all().order_by('-created_at')

        # Filter by village_id (supports multiple comma-separated IDs)
        if village_id:
            village_id_list = [vid.strip() for vid in village_id.split(',')]
            try:
                village_id_list = [int(vid) for vid in village_id_list]  # Convert to integers
                queryset = queryset.filter(village__id__in=village_id_list)
            except ValueError:
                # Handle invalid village_id format gracefully
                pass

        # Filter by tehsil_id (supports multiple comma-separated IDs)
        if tehsil_id:
            tehsil_id_list = [tid.strip() for tid in tehsil_id.split(',')]
            try:
                tehsil_id_list = [int(tid) for tid in tehsil_id_list]  # Convert to integers
                queryset = queryset.filter(village__tehsil__id__in=tehsil_id_list)
            except ValueError:
                # Handle invalid tehsil_id format gracefully
                pass

        # Filter by city_id (supports multiple comma-separated IDs)
        if city_id:
            city_id_list = [cid.strip() for cid in city_id.split(',')]
            try:
                city_id_list = [int(cid) for cid in city_id_list]  # Convert to integers
                queryset = queryset.filter(village__tehsil__city__id__in=city_id_list)
            except ValueError:
                # Handle invalid city_id format gracefully
                pass


        # Search filter for farmer name, father name, mobile number, email, and location fields
        if search_query:
            search_terms = [term.strip() for term in search_query.split(',') if term.strip()]
            if search_terms:
                q = models.Q()
                for term in search_terms:
                    q |= (
                        models.Q(form__farmerName__icontains=term) |
                        models.Q(form__fatherName__icontains=term) |
                        models.Q(form__mobileNumber__icontains=term) |
                        models.Q(form__email__icontains=term) |
                        models.Q(village__pin_code__icontains=term) |
                        models.Q(village__display_name__icontains=term) |
                        models.Q(village__tehsil__name__icontains=term) |
                        models.Q(village__tehsil__city__name__icontains=term) 
                    )
                queryset = queryset.filter(q)
            else:
                # If search_query is empty or only contains invalid terms, return empty queryset
                queryset = queryset.none()

        # Search filter for farmer name (supports multiple comma-separated values with icontains)
        if farmer_name:
            farmer_name_list = [name.strip() for name in farmer_name.split(',') if name.strip()]
            if farmer_name_list:
                q = models.Q()
                for name in farmer_name_list:
                    q |= models.Q(form__farmerName__icontains=name)
                queryset = queryset.filter(q)
            else:
                # If farmer_name is empty or invalid, return empty queryset
                queryset = queryset.none()

        # Filter by house_status (supports multiple comma-separated values with icontains)
        if house_status:
            house_status_list = [status.strip() for status in house_status.split(',') if status.strip()]
            if house_status_list:
                q = models.Q()
                for status in house_status_list:
                    q |= models.Q(form__houseStatus__icontains=status)
                queryset = queryset.filter(q)
            else:
                # If house_status is empty or invalid, return empty queryset
                queryset = queryset.none()

        # Filter by estimatedCropLoss range (e.g., estimatedCropLoss=1000-20000)
        if estimated_crop_loss:
            try:
                min_loss, max_loss = map(float, estimated_crop_loss.split('-'))
                queryset = queryset.filter(
                    form__estimatedCropLoss__gte=min_loss,
                    form__estimatedCropLoss__lte=max_loss
                )
            except ValueError:
                # Handle invalid range format by returning empty queryset
                queryset = queryset.none()

        # Apply pagination
        paginator = self.pagination_class()
        page = paginator.paginate_queryset(queryset, request)
        data = []

        for person_data in page:
            verification_data = None
            if person_data.is_single_user:
                try:
                    verified_by = person_data.form.verified_by
                    verification_data = [{
                        'verified_by_data': [{
                            'verification_id': verified_by.id,
                            'verification_image': urljoin(
                                request.build_absolute_uri('/'),
                                verified_by.verification_image.url
                            ) if verified_by.verification_image else None,
                            'surveyor_name': verified_by.surveyor_name,
                            'surveyor_mobile': verified_by.surveyor_mobile,
                            'date': verified_by.date,
                            'is_verified': verified_by.is_verified,
                        }]
                    }]
                except AttributeError:
                    verification_data = None

            data.append({
                'person_id': person_data.id,
                'is_active': person_data.is_active,
                'is_single_user': person_data.is_single_user,
                'created_at':person_data.created_at,
                "village_data": [{
                    'village_id': person_data.village.id,
                    'display_name': person_data.village.display_name,
                    'pin_code': person_data.village.pin_code,
                    'longitude': person_data.village.longitude,
                    'latitude': person_data.village.latitude,
                    "tehsil_data": [{
                        'tehsil_id': person_data.village.tehsil.id,
                        'name': person_data.village.tehsil.name,
                        'city_data': [{
                            'city_id': person_data.village.tehsil.city.id,
                            'name': person_data.village.tehsil.city.name,
                            'state_data': [{
                                'state_id': person_data.village.tehsil.city.state.id,
                                'name': person_data.village.tehsil.city.state.name
                            }]
                        }]
                    }]
                }],
                "form_data": [{
                    "form_id": person_data.form.id,
                    'farmer_name': person_data.form.farmerName,
                    'farmer_image': urljoin(
                        request.build_absolute_uri('/'),
                        person_data.form.farmer_image.url
                    ) if person_data.form.farmer_image else None,
                    'mobileNumber': person_data.form.mobileNumber,
                    'fatherName': person_data.form.fatherName,
                    'email': person_data.form.email,
                    'houseStatus': person_data.form.houseStatus,
                    'totalLandOwned': person_data.form.totalLandOwned,
                    'landAffected': person_data.form.landAffected,
                    'cropsPlanted': person_data.form.cropsPlanted,
                    'cropsLost': person_data.form.cropsLost,
                    'estimatedCropLoss': person_data.form.estimatedCropLoss,
                    'tractorLeveling': person_data.form.tractorLeveling,
                    'manureFertilizer': person_data.form.manureFertilizer,
                    'seedsRequired': person_data.form.seedsRequired,
                    'fertilizersPesticides': person_data.form.fertilizersPesticides,
                    'laborRequirement': person_data.form.laborRequirement,
                    'irrigationRepair': person_data.form.irrigationRepair,
                    'livestockDamage': person_data.form.livestockDamage,
                    'householdNeeds': person_data.form.householdNeeds,
                    'housingRepair': person_data.form.housingRepair,
                    'otherSupport': person_data.form.otherSupport,
                    'additionalNotes': person_data.form.additionalNotes,
                    'amount_needed': person_data.form.amount_needed if person_data.form.amount_needed else None,
                    'amount_received': person_data.form.amount_received if person_data.form.amount_received else None
                }],
                'verification_data': verification_data
            })

        return paginator.get_paginated_response(data)








class ExportLocationDataToExcelAPIView(APIView):
    def get(self, request):
        try:
            # Create workbook
            wb = Workbook()
            
            # Sheet 1: States
            ws_state = wb.create_sheet(title="States")
            state_headers = ['State ID', 'State Name']
            for col, header in enumerate(state_headers, 1):
                cell = ws_state.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
            
            # Populate State data
            states = State.objects.all()
            for row, state in enumerate(states, 2):
                ws_state.cell(row=row, column=1, value=state.id)
                ws_state.cell(row=row, column=2, value=state.name)
            
            # Sheet 2: Cities
            ws_city = wb.create_sheet(title="Cities")
            city_headers = ['City ID', 'City Name']
            for col, header in enumerate(city_headers, 1):
                cell = ws_city.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
            
            # Populate City data
            cities = City.objects.all()
            for row, city in enumerate(cities, 2):
                ws_city.cell(row=row, column=1, value=city.id)
                ws_city.cell(row=row, column=2, value=city.name)
            
            # Sheet 3: Tehsils
            ws_tehsil = wb.create_sheet(title="Tehsils")
            tehsil_headers = ['Tehsil ID', 'Tehsil Name']
            for col, header in enumerate(tehsil_headers, 1):
                cell = ws_tehsil.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
            
            # Populate Tehsil data
            tehsils = Tehsil.objects.all()
            for row, tehsil in enumerate(tehsils, 2):
                ws_tehsil.cell(row=row, column=1, value=tehsil.id)
                ws_tehsil.cell(row=row, column=2, value=tehsil.name)
            
            # Sheet 4: Villages
            ws_village = wb.create_sheet(title="Villages")
            village_headers = ['Village ID', 'Village Name']
            for col, header in enumerate(village_headers, 1):
                cell = ws_village.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
            
            # Populate Village data
            villages = Village.objects.all()
            for row, village in enumerate(villages, 2):
                ws_village.cell(row=row, column=1, value=village.id)
                ws_village.cell(row=row, column=2, value=village.display_name)
            
            # Remove default sheet created by Workbook
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
            
            # Auto-fit column widths for all sheets
            for ws in wb:
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
            
            # Enable sheet protection for all sheets (lock all cells)
            for ws in wb:
                ws.protection = SheetProtection(
                    sheet=True,
                    objects=True,
                    scenarios=True,
                    formatCells=True,
                    formatColumns=True,
                    formatRows=True,
                    insertRows=True,
                    insertColumns=True,
                    deleteRows=True,
                    deleteColumns=True
                )
            
            # Create excel_reports directory if it doesn't exist
            excel_dir = os.path.join(settings.MEDIA_ROOT, 'excel_reports')
            os.makedirs(excel_dir, exist_ok=True)
            
            # Generate unique filename with timestamp
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f'location_data_{timestamp}.xlsx'
            file_path = os.path.join(excel_dir, filename)
            
            # Save the workbook to the file
            wb.save(file_path)
            
            # Construct absolute URL
            relative_url = f'{settings.MEDIA_URL}excel_reports/{filename}'
            absolute_url = request.build_absolute_uri(relative_url)
            
            # Return JSON response with the absolute URL
            response_data = {
                'message': 'Location data Excel file generated successfully',
                'download_url': absolute_url,
                'filename': filename
            }
            
            return Response(response_data, status=status.HTTP_200_OK)
        
        except Exception as e:
            import traceback
            print("❌ ERROR: Failed to generate Excel file")
            print(traceback.format_exc())
            return Response(
                {'error': f'Failed to generate Excel file: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )



    





from pathlib import Path
import random
from django.http import FileResponse


class GenerateDummyDataExcelAPIView(APIView):
    def get(self, request, *args, **kwargs):
        try:
            # Configuration
            output_dir = Path(settings.MEDIA_ROOT) / 'excel_reports'
            output_dir.mkdir(exist_ok=True)
            output_file = output_dir / 'dummy_farmers_data_500.xlsx'

            # Dummy data lists
            first_names = [
                'Gurpreet', 'Jaswinder', 'Baljit', 'Amarjeet', 'Manpreet',
                'Harpreet', 'Sukhwinder', 'Daljit', 'Karan', 'Ravinder',
                'Simran', 'Ranjit', 'Kuldeep', 'Sandeep', 'Navjot',
                'Hardeep', 'Paramjit', 'Jagjit', 'Satnam', 'Ajay',
                'Pardeep', 'Rajinder', 'Mandeep', 'Surinder', 'Balwinder',
                'Gagan', 'Harpal', 'Jaspal', 'Kamaljit', 'Lakhwinder',
                'Narinder', 'Prabhjot', 'Rajwinder', 'Sukhdev', 'Taranjit',
                'Amritpal', 'Charanjit', 'Darshan', 'Gurmail', 'Harjinder',
                'Jagmohan', 'Kashmir', 'Kirpal', 'Manjit', 'Nirmal',
                'Raghbir', 'Satpal', 'Shamsher', 'Surjit', 'Tejinder'
            ]
            last_names = ['Singh', 'Kaur', 'Sharma']
            father_names = [
                'Harbans Singh', 'Gurcharan Singh', 'Baldev Singh',
                'Sohan Singh', 'Amrik Singh'
            ]
            surveyor_names = [
                'Amarjeet Singh', 'Manpreet Kaur', 'Ranjit Singh',
                'Simerpreet Kaur', 'Kuldeep Singh'
            ]
            house_statuses = ['Partially Damaged', 'No Damage', 'Fully Damaged']
            crop_losses = ['10,000 - 25,000', '25,000 - 50,000', '50,000 - 100,000']
            yes_no_options = ['Yes', 'No', 'Not Required']
            land_values = ['5', '10', '15', '20']
            crop_values = ['50', '100', '200', '300']

            # Headers
            headers = [
                'farmerName', 'fatherName', 'mobileNumber', 'email', 'state_id',
                'city_id', 'tehsil_id', 'village_id', 'houseStatus',
                'totalLandOwned', 'landAffected', 'cropsPlanted', 'cropsLost',
                'estimatedCropLoss', 'tractorLeveling', 'manureFertilizer',
                'seedsRequired', 'fertilizersPesticides', 'laborRequirement',
                'irrigationRepair', 'livestockDamage', 'householdNeeds',
                'housingRepair', 'otherSupport', 'surveyorName', 'AmountNeeded'
            ]

            # Load valid location hierarchy from models
            def load_location_hierarchy():
                with transaction.atomic():
                    states = list(State.objects.all())
                    if not states:
                        raise ValueError("No states found in the database")
                    
                    state = states[0]  # Assuming only Punjab for now
                    cities = list(City.objects.filter(state=state))
                    if not cities:
                        raise ValueError("No cities found for the state")
                    
                    village_list = []
                    for city in cities:
                        tehsils = list(Tehsil.objects.filter(city=city))
                        if not tehsils:
                            continue
                        villages = list(Village.objects.filter(tehsil__in=tehsils))
                        for village in villages:
                            village_list.append((village.id, state.id, city.id, village.tehsil.id))
                    
                    if not village_list:
                        raise ValueError("No valid village hierarchy found")
                    
                    return village_list

            village_list = load_location_hierarchy()

            # Generate rows
            def generate_row(village_idx: int):
                village_id, state_id, city_id, tehsil_id = village_list[village_idx % len(village_list)]

                first = random.choice(first_names)
                last = random.choice(last_names)
                farmer_name = f"{first} {last}"
                email = f"{first.lower()}.{last.lower()}{random.randint(100, 999)}@gmail.com" if random.random() > 0.3 else ''

                total_land = random.choice(land_values)
                affected_land = str(random.randint(1, int(float(total_land))))
                planted = random.choice(crop_values)
                lost = str(random.randint(10, int(float(planted))))
                amount_needed = random.randint(5000, 50000)

                return [
                    farmer_name,
                    random.choice(father_names),
                    f"{random.randint(7,9)}{random.randint(10000000,99999999)}",
                    email,
                    state_id,
                    city_id,
                    tehsil_id,
                    village_id,
                    '',  # pincode – left empty
                    random.choice(house_statuses),
                    total_land,
                    affected_land,
                    planted,
                    lost,
                    random.choice(crop_losses),
                    random.choice(yes_no_options),
                    random.choice(yes_no_options),
                    random.choice(yes_no_options),
                    random.choice(yes_no_options),
                    random.choice(yes_no_options),
                    random.choice(yes_no_options),
                    random.choice(yes_no_options),
                    random.choice(yes_no_options),
                    random.choice(yes_no_options),
                    random.choice(yes_no_options),
                    random.choice(surveyor_names),
                    amount_needed
                ]

            rows = [generate_row(i) for i in range(500)]

            # Create Excel file
            wb = Workbook()
            ws = wb.active
            ws.title = 'Farmers Data'

            # Headers
            for col, hdr in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=hdr)
                cell.font = Font(bold=True)
                cell.protection = Protection(locked=True)

            # Data
            for r_idx, row_data in enumerate(rows, start=2):
                for c_idx, value in enumerate(row_data, start=1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)
                    cell.protection = Protection(locked=False)

            # Auto-fit columns
            for column in ws.columns:
                max_len = 0
                col_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_len:
                            max_len = len(str(cell.value))
                    except Exception:
                        pass
                adjusted_width = min(max_len + 2, 50)
                ws.column_dimensions[col_letter].width = adjusted_width

            # Protect the sheet
            ws.protection = SheetProtection(sheet=True, objects=False, scenarios=False)

            # Save the file
            wb.save(output_file)

            # Return the file as a response
            response = FileResponse(
                open(output_file, 'rb'),
                as_attachment=True,
                filename='dummy_farmers_data_500.xlsx'
            )
            return response

        except ValueError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({'error': f'Server error: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)









class EditFarmerDataAPIViewAdmin(APIView):

    def patch(self, request):
        print("=" * 60)
        print("EDIT FARMER DATA ADMIN APIVIEW - PATCH REQUEST STARTED")
        print("=" * 60)

        try:
            # ------------------------------------------------------------------
            # 1. Get the PersonData instance that ties everything together
            # ------------------------------------------------------------------
            person_data_id = request.data.get("person_data_id")
            if not person_data_id:
                raise ValueError("person_data_id is required")

            person_data = PersonData.objects.select_related(
                'form__verified_by', 'village__tehsil__city__state'
            ).get(id=person_data_id)

            # ------------------------------------------------------------------
            # 2. Parse JSON string (same format you used for create)
            # ------------------------------------------------------------------
            data_str = request.data.get("data")
            if not data_str:
                raise ValueError("JSON data field is required")

            data = json.loads(data_str)  # single dict, not a list

            # ------------------------------------------------------------------
            # 3. Start transaction
            # ------------------------------------------------------------------
            with transaction.atomic():
                print("Transaction STARTED")

                # ------------------- LOCATION HIERARCHY -------------------
                state_id   = data.get('state_id')
                city_id    = data.get('city_id')
                tehsil_id  = data.get('tehsil_id')
                village_id = data.get('village_id')

                # Validate hierarchy (same logic as create)
                state   = State.objects.get(id=state_id)
                city    = City.objects.get(id=city_id, state=state)
                tehsil  = Tehsil.objects.get(id=tehsil_id, city=city)
                village = Village.objects.get(id=village_id, tehsil=tehsil)

                # ------------------- VERIFIED BY -------------------------
                verified_data = data.get('verifiedBy', [{}])[0]

                # Optional image replacement
                new_verification_image = request.FILES.get('verification_image')
                verified_by = person_data.form.verified_by

                verified_by.surveyor_name = verified_data.get('surveyorName', verified_by.surveyor_name)
                verified_by.surveyor_mobile = verified_data.get('surveyorMobile', verified_by.surveyor_mobile or '')

                date_str = verified_data.get('date')
                if date_str:
                    parsed = parse_date(date_str)
                    if not parsed:
                        raise ValueError(f"Invalid date format: {date_str}")
                    verified_by.date = parsed

                if new_verification_image:
                    # Delete old file if you want (optional)
                    if verified_by.verification_image:
                        verified_by.verification_image.delete(save=False)
                    verified_by.verification_image = new_verification_image

                verified_by.is_verified = verified_data.get('is_verified', verified_by.is_verified)
                verified_by.save()
                print(f"VerifiedBy UPDATED - ID: {verified_by.id}")

                # ------------------- FARMER FORM -------------------------
                farmer_form = person_data.form
                new_farmer_image = request.FILES.get('farmer_image')

                if new_farmer_image:
                    if farmer_form.farmer_image:
                        farmer_form.farmer_image.delete(save=False)
                    farmer_form.farmer_image = new_farmer_image

                # Update all editable fields (skip User-related)
                field_map = {
                    'farmerName': 'farmerName',
                    'fatherName': 'fatherName',
                    'mobileNumber': 'mobileNumber',
                    'email': 'email',                     # allowed – not part of User model
                    'houseStatus': 'houseStatus',
                    'totalLandOwned': 'totalLandOwned',
                    'landAffected': 'landAffected',
                    'cropsPlanted': 'cropsPlanted',
                    'cropsLost': 'cropsLost',
                    'estimatedCropLoss': 'estimatedCropLoss',
                    'tractorLeveling': 'tractorLeveling',
                    'manureFertilizer': 'manureFertilizer',
                    'seedsRequired': 'seedsRequired',
                    'fertilizersPesticides': 'fertilizersPesticides',
                    'laborRequirement': 'laborRequirement',
                    'irrigationRepair': 'irrigationRepair',
                    'livestockDamage': 'livestockDamage',
                    'householdNeeds': 'householdNeeds',
                    'housingRepair': 'housingRepair',
                    'otherSupport': 'otherSupport',
                    'additionalNotes': 'additionalNotes',
                    'amount_needed': 'amount_needed',
                    'amount_received': 'amount_received',
                    'is_active': 'is_active',
                }

                for json_key, model_field in field_map.items():
                    value = data.get(json_key)
                    if value is not None:
                        setattr(farmer_form, model_field, value)

                farmer_form.save()
                print(f"FarmerForm UPDATED - ID: {farmer_form.id}")

                # ------------------- PERSON DATA -------------------------
                # Only update village & flags (user stays untouched)
                person_data.village = village
                person_data.is_active = data.get('is_active', person_data.is_active)
                person_data.is_single_user = data.get('is_single_user', person_data.is_single_user)
                person_data.save()
                print(f"PersonData UPDATED - ID: {person_data.id}")

                # ------------------- RESPONSE -------------------------
                response_payload = {
                    "message": "Farmer data updated successfully",
                    "person_data_id": person_data.id,
                    "farmer_form_id": farmer_form.id,
                    "village": village.display_name,
                    "updated_at": farmer_form.updated_at.isoformat(),
                }

                print("Transaction COMMITTED")
                return Response(response_payload, status=status.HTTP_200_OK)

        # ------------------------------------------------------------------
        # ERROR HANDLING (same style as your POST view)
        # ------------------------------------------------------------------
        except json.JSONDecodeError as e:
            print(f"JSON DECODE ERROR: {e}")
            return Response(
                {"error": f"Invalid JSON in data field: {str(e)}"},
                status=status.HTTP_400_BAD_REQUEST,
            )
        except ValueError as e:
            print(f"VALUE ERROR: {e}")
            print("Transaction ROLLEDBACK")
            return Response(
                {"error": f"Validation error: {str(e)}"},
                status=status.HTTP_400_BAD_REQUEST,
            )
        except PersonData.DoesNotExist:
            return Response(
                {"error": "PersonData not found"},
                status=status.HTTP_404_NOT_FOUND,
            )
        except Exception as e:
            print(f"GENERAL EXCEPTION: {e}")
            traceback.print_exc()
            print("Transaction ROLLEDBACK")
            return Response(
                {"error": f"Server error: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )
        finally:
            print("=" * 60)
            print("EDIT FARMER DATA ADMIN APIVIEW - PATCH REQUEST FINISHED")
            print("=" * 60)





class DeleteFarmerDataAPIViewAdmin(APIView):
    """
    DELETE /api/farmer/delete/
    {
        "person_data_id": 57
    }
    """

    def post(self, request):
        print("=" * 60)
        print("DELETE FARMER DATA ADMIN APIVIEW - DELETE REQUEST STARTED")
        print("=" * 60)

        person_data_id = request.data.get("person_data_id")
        if not person_data_id:
            print("ERROR: person_data_id is required")
            return Response(
                {"error": "person_data_id is required"},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            # Use select_related to fetch all in one query
            person_data = PersonData.objects.select_related(
                'form__verified_by'
            ).get(id=person_data_id)

            print(f"Found PersonData ID: {person_data.id}")
            print(f"Farmer: {person_data.form.farmerName}")
            print(f"Mobile: {person_data.form.mobileNumber}")
            print(f"Village: {person_data.village.display_name}")

            with transaction.atomic():
                print("Transaction STARTED")

                # 1. Store references
                farmer_form = person_data.form
                verified_by = farmer_form.verified_by

                # 2. Delete images first (to free storage)
                if farmer_form.farmer_image:
                    print(f"Deleting farmer_image: {farmer_form.farmer_image.path}")
                    farmer_form.farmer_image.delete(save=False)

                if verified_by.verification_image:
                    print(f"Deleting verification_image: {verified_by.verification_image.path}")
                    verified_by.verification_image.delete(save=False)

                # 3. Delete in correct order (FK constraints)
                person_data.delete()        # Deletes PersonData
                print(f"PersonData DELETED - ID: {person_data.id}")

                farmer_form.delete()        # Deletes FarmerForm
                print(f"FarmerForm DELETED - ID: {farmer_form.id}")

                verified_by.delete()        # Deletes VerifiedBy
                print(f"VerifiedBy DELETED - ID: {verified_by.id}")

                print("Transaction COMMITTED")

            return Response({
                "message": "Farmer data deleted successfully",
                "person_data_id": person_data_id,
                "farmer_name": farmer_form.farmerName,
                "user_id": person_data.user.id,  # User is NOT deleted
                "user_kept": True
            }, status=status.HTTP_200_OK)

        except PersonData.DoesNotExist:
            print("ERROR: PersonData not found")
            return Response(
                {"error": "PersonData not found"},
                status=status.HTTP_404_NOT_FOUND
            )

        except Exception as e:
            print(f"GENERAL EXCEPTION: {str(e)}")
            traceback.print_exc()
            print("Transaction ROLLEDBACK")
            return Response(
                {"error": f"Server error: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

        finally:
            print("=" * 60)
            print("DELETE FARMER DATA ADMIN APIVIEW - REQUEST FINISHED")
            print("=" * 60)





class GetFarmerDataAPIViewAdmin(APIView):
    """
    GET /api/farmer/get/
    Query: ?person_data_id=57
    OR
    POST: { "person_data_id": 57 }
    """

    def get(self, request):
        print("=" * 60)
        print("GET FARMER DATA ADMIN APIVIEW - REQUEST STARTED")
        print("=" * 60)

        # Accept from query param OR JSON body
        person_data_id = request.query_params.get("person_data_id") 

        if not person_data_id:
            print("ERROR: person_data_id is required")
            return Response(
                {"error": "person_data_id is required"},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            person_data = PersonData.objects.select_related(
                'form__verified_by',
                'village__tehsil__city__state',
                'user'
            ).get(id=person_data_id)

            farmer_form = person_data.form
            verified_by = farmer_form.verified_by

            # Build location chain
            village = person_data.village
            tehsil = village.tehsil
            city = tehsil.city
            state = city.state

            # Image URLs
            farmer_image_url = (
                request.build_absolute_uri(farmer_form.farmer_image.url)
                if farmer_form.farmer_image else None
            )
            verification_image_url = (
                request.build_absolute_uri(verified_by.verification_image.url)
                if verified_by.verification_image else None
            )

            # Response payload
            data = {
                "person_data_id": person_data.id,
                "user_id": person_data.user.id,  # Only ID, no password/email

                # Location
                "state_id": state.id,
                "state_name": state.name,
                "city_id": city.id,
                "city_name": city.name,
                "tehsil_id": tehsil.id,
                "tehsil_name": tehsil.name,
                "village_id": village.id,
                "village_name": village.display_name,

                # VerifiedBy
                "verifiedBy": {
                    "surveyorName": verified_by.surveyor_name,
                    "surveyorMobile": verified_by.surveyor_mobile or "",
                    "date": verified_by.date.isoformat() if verified_by.date else None,
                    "is_verified": verified_by.is_verified,
                    "verification_image": verification_image_url
                },

                # FarmerForm fields
                "farmerName": farmer_form.farmerName,
                "fatherName": farmer_form.fatherName,
                "mobileNumber": farmer_form.mobileNumber,
                "email": farmer_form.email,
                "houseStatus": farmer_form.houseStatus,
                "totalLandOwned": farmer_form.totalLandOwned,
                "landAffected": farmer_form.landAffected,
                "cropsPlanted": farmer_form.cropsPlanted,
                "cropsLost": farmer_form.cropsLost,
                "estimatedCropLoss": farmer_form.estimatedCropLoss,
                "tractorLeveling": farmer_form.tractorLeveling,
                "manureFertilizer": farmer_form.manureFertilizer,
                "seedsRequired": farmer_form.seedsRequired,
                "fertilizersPesticides": farmer_form.fertilizersPesticides,
                "laborRequirement": farmer_form.laborRequirement,
                "irrigationRepair": farmer_form.irrigationRepair,
                "livestockDamage": farmer_form.livestockDamage,
                "householdNeeds": farmer_form.householdNeeds,
                "housingRepair": farmer_form.housingRepair,
                "otherSupport": farmer_form.otherSupport,
                "additionalNotes": farmer_form.additionalNotes,
                "amount_needed": farmer_form.amount_needed,
                "amount_received": farmer_form.amount_received,
                "is_active": farmer_form.is_active,
                "farmer_image": farmer_image_url,

                # PersonData flags
                "person_is_active": person_data.is_active,
                "is_single_user": person_data.is_single_user,

                # Timestamps
                "created_at": farmer_form.created_at.isoformat(),
                "updated_at": farmer_form.updated_at.isoformat(),
            }

            print(f"SUCCESS: Data retrieved for PersonData ID: {person_data_id}")
            print(f"Farmer: {data['farmerName']}")
            print(f"Village: {data['village_name']}")

            return Response({
                "message": "Farmer data retrieved successfully",
                "data": data
            }, status=status.HTTP_200_OK)

        except PersonData.DoesNotExist:
            print("ERROR: PersonData not found")
            return Response(
                {"error": "PersonData not found"},
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            print(f"GENERAL EXCEPTION: {str(e)}")
            import traceback
            traceback.print_exc()
            return Response(
                {"error": f"Server error: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        finally:
            print("=" * 60)
            print("GET FARMER DATA ADMIN APIVIEW - REQUEST FINISHED")
            print("=" * 60)